"""
Daily Runner — Automated Results Pipeline
--------------------------------------------
Orchestrates the entire daily workflow:

  1. BSE Board Meetings: Scrape next 45 days → Board_Meetings.csv → Board_Meetings_Report.xlsx
  2. Results: Detect missed dates → scrape each → Results_By_Date.xlsx, Results_By_Sector.xlsx
  3. Public Website: Regenerate unified HTML dashboard with all data
  4. (Optional) Git push to GitHub Pages

Smart missed-date detection:
  - Reads Results_By_Date.xlsx to find last processed date
  - Processes all dates from last_processed+1 to yesterday (inclusive)
  - Handles weekends, holidays, and any gaps

Usage:
  py daily_runner.py                  # auto-detect missed dates + BSE scrape
  py daily_runner.py --skip-bse      # skip BSE scrape, only results + dashboard
  py daily_runner.py --force-all     # force re-scrape from April 1
"""

import os
import sys
import subprocess
import time
import json
import shutil
from datetime import datetime, timedelta
from pathlib import Path

DIR = os.path.dirname(os.path.abspath(__file__))
PYTHON = sys.executable

# File paths
DATE_WB_PATH = os.path.join(DIR, "Results_By_Date.xlsx")
SECTOR_WB_PATH = os.path.join(DIR, "Results_By_Sector.xlsx")
DAILY_DASHBOARD = os.path.join(DIR, "Daily_Dashboard.html")
SECTOR_DASHBOARD = os.path.join(DIR, "Sector_Dashboard.html")
BOARD_CSV = os.path.join(DIR, "Board_Meetings.csv")
BOARD_REPORT = os.path.join(DIR, "Board_Meetings_Report.xlsx")
PUBLIC_SITE_DIR = os.path.dirname(DIR)
RUN_LOG = os.path.join(DIR, "last_run.json")

# Season start — don't go back further than this
SEASON_START = datetime(2026, 4, 1)


def log(msg, level="INFO"):
    ts = datetime.now().strftime("%H:%M:%S")
    print(f"  [{ts}] {level}: {msg}")


def get_last_processed_date():
    """
    Read Results_By_Date.xlsx sheet names to find the most recent processed date.
    Sheet names are formatted as 'DD-Mon-YYYY'.
    """
    if not os.path.exists(DATE_WB_PATH):
        return None

    try:
        from openpyxl import load_workbook
        wb = load_workbook(DATE_WB_PATH, read_only=True)
        dates = []
        for name in wb.sheetnames:
            for fmt in ("%d-%b-%Y", "%d-%B-%Y"):
                try:
                    dates.append(datetime.strptime(name, fmt))
                    break
                except ValueError:
                    continue
        wb.close()

        if dates:
            latest = max(dates)
            log(f"Last processed date in workbook: {latest.strftime('%d %b %Y')}")
            return latest
    except Exception as e:
        log(f"Error reading workbook: {e}", "WARN")

    return None


def get_dates_to_process():
    """
    Determine which dates need processing.
    Goes back to find first unprocessed date, up to SEASON_START.
    """
    yesterday = datetime.today() - timedelta(days=1)
    # Don't process today (results won't be out yet)
    end_date = yesterday

    last_processed = get_last_processed_date()

    if last_processed:
        # Start from the day after last processed
        start_date = last_processed + timedelta(days=1)
    else:
        # No workbook exists — start from season start
        start_date = SEASON_START

    if start_date > end_date:
        log("All dates up to yesterday are already processed.")
        return []

    dates = []
    d = start_date
    while d <= end_date:
        dates.append(d)
        d += timedelta(days=1)

    log(f"Dates to process: {len(dates)} "
        f"({dates[0].strftime('%d %b')} → {dates[-1].strftime('%d %b')})")
    return dates


def run_bse_scraper():
    """Step 1: Scrape BSE board meetings for next 45 days."""
    log("=" * 50)
    log("STEP 1: Scraping BSE Board Meetings (next 45 days)")
    log("=" * 50)

    script = os.path.join(DIR, "bse_scraper.py")
    if not os.path.exists(script):
        log("bse_scraper.py not found — skipping BSE scrape", "WARN")
        return False

    try:
        result = subprocess.run(
            [PYTHON, script, "45"],
            cwd=DIR, capture_output=True, text=True, timeout=300
        )
        print(result.stdout)
        if result.stderr:
            print(result.stderr)
        return result.returncode == 0
    except Exception as e:
        log(f"BSE scraper failed: {e}", "ERROR")
        return False


def run_indices_scraper():
    """Step 1.1: Scrape Nifty indices constituents."""
    log("=" * 50)
    log("STEP 1.1: Scraping Screener Indices")
    log("=" * 50)

    script = os.path.join(DIR, "update_indices.py")
    if not os.path.exists(script):
        log("update_indices.py not found — skipping indices scrape", "WARN")
        return False

    try:
        result = subprocess.run(
            [PYTHON, script],
            cwd=DIR, capture_output=True, text=True, timeout=300
        )
        print(result.stdout)
        if result.stderr:
            print(result.stderr)
        return result.returncode == 0
    except Exception as e:
        log(f"Indices scraper failed: {e}", "ERROR")
        return False



def run_board_processor():
    """Step 2: Run board.py to enrich Board_Meetings.csv with MCap & sectors."""
    log("=" * 50)
    log("STEP 2: Processing Board Meetings Report")
    log("=" * 50)

    script = os.path.join(DIR, "board.py")
    if not os.path.exists(script):
        log("board.py not found — skipping", "WARN")
        return False
    if not os.path.exists(BOARD_CSV):
        log("Board_Meetings.csv not found — skipping", "WARN")
        return False

    try:
        result = subprocess.run(
            [PYTHON, script],
            cwd=DIR, capture_output=True, text=True, timeout=600
        )
        print(result.stdout)
        if result.stderr:
            print(result.stderr)
        return result.returncode == 0
    except Exception as e:
        log(f"Board processor failed: {e}", "ERROR")
        return False


def run_sorter(dates):
    """Step 3: Run sorter.py for missed dates."""
    if not dates:
        log("No dates to process — skipping sorter.")
        return True

    log("=" * 50)
    log(f"STEP 3: Running Results Sorter ({len(dates)} dates)")
    log("=" * 50)

    script = os.path.join(DIR, "sorter.py")
    if not os.path.exists(script):
        log("sorter.py not found — skipping", "WARN")
        return False

    start_str = dates[0].strftime("%Y-%m-%d")
    end_str = dates[-1].strftime("%Y-%m-%d")

    try:
        result = subprocess.run(
            [PYTHON, script, start_str, end_str],
            cwd=DIR, capture_output=True, text=True, timeout=3600
        )
        print(result.stdout)
        if result.stderr:
            print(result.stderr)
        return result.returncode == 0
    except Exception as e:
        log(f"Sorter failed: {e}", "ERROR")
        return False


def build_public_site():
    """Step 4: Generate public website."""
    log("=" * 50)
    log("STEP 4: Building Public Website")
    log("=" * 50)

    os.makedirs(PUBLIC_SITE_DIR, exist_ok=True)

    # Read data from the generated dashboards
    daily_data = _read_daily_dashboard_data()
    sector_data = _read_sector_dashboard_data()
    board_data = _read_board_meetings_data()

    # Generate unified site
    _generate_unified_html(daily_data, sector_data, board_data)

    log(f"[OK] Public site generated: {PUBLIC_SITE_DIR}")
    return True


def _read_daily_dashboard_data():
    """
    Read ALL date sheets from Results_By_Date.xlsx and return records/dates/sectors.

    Previously read from Daily_Dashboard.html which gets overwritten by every
    sorter.py run, causing all but the last date to vanish from the dashboard.
    """
    if not os.path.exists(DATE_WB_PATH):
        return {"records": [], "dates": [], "sectors": []}

    from openpyxl import load_workbook

    ALL_COLS = [
        "Result Date", "Company Name", "Sector", "Industry Group", "Subsector", "Indices", "Market Cap (Cr)", "Price", "PE", "Screener Ticker",
        "Quarter",
        "Sales (Cr)", "EBITDA (Cr)", "Net Profit (Cr)", "EPS (Rs)",
        "EBITDA Margin%", "PAT Margin%",
        "Sales QoQ%", "EBITDA QoQ%", "NP QoQ%", "EPS QoQ%",
        "EBITDA Margin QoQ pp", "PAT Margin QoQ pp",
        "Sales YoY Q%", "EBITDA YoY Q%", "NP YoY Q%", "EPS YoY Q%",
        "EBITDA Margin YoY pp", "PAT Margin YoY pp",
        "Prev Qtr Sales", "Prev Qtr EBITDA", "Prev Qtr NP", "Prev Qtr EPS",
        "LY Qtr Sales", "LY Qtr EBITDA", "LY Qtr NP", "LY Qtr EPS",
        "FY Sales (Cr)", "FY EBITDA (Cr)", "FY NP (Cr)", "FY EPS (Rs)",
        "FY EBITDA Margin%", "FY PAT Margin%",
        "FY Sales YoY%", "FY EBITDA YoY%", "FY NP YoY%", "FY EPS YoY%",
        "FY EBITDA Margin YoY pp", "FY PAT Margin YoY pp",
        "Prev FY Sales", "Prev FY EBITDA", "Prev FY NP", "Prev FY EPS",
        "Screener URL",
    ]
    EXPOSE_COLS = [
        "Result Date", "Company Name", "Sector", "Industry Group", "Subsector", "Indices", "Market Cap (Cr)", "Quarter", "Screener Ticker",
        "Sales (Cr)", "EBITDA (Cr)", "Net Profit (Cr)", "EPS (Rs)",
        "EBITDA Margin%", "PAT Margin%",
        "Sales YoY Q%", "EBITDA YoY Q%", "NP YoY Q%", "EPS YoY Q%",
        "EBITDA Margin YoY pp", "PAT Margin YoY pp",
        "Sales QoQ%", "EBITDA QoQ%", "NP QoQ%", "EPS QoQ%",
        "EBITDA Margin QoQ pp", "PAT Margin QoQ pp",
        "FY Sales YoY%", "FY EBITDA YoY%", "FY NP YoY%", "FY EPS YoY%",
        "FY EBITDA Margin YoY pp", "FY PAT Margin YoY pp",
        "Screener URL",
    ]
    STR_COLS = {"Result Date", "Company Name", "Sector", "Industry Group", "Subsector", "Indices", "Quarter", "Screener Ticker"}
    SKIP_PREFIXES = ("Result Date", "Results  ", "\u25b6", "Sorted")

    def safe_num(v):
        if v is None or v == "":
            return None
        try:
            return float(v)
        except (TypeError, ValueError):
            return str(v).strip() if v else None

    def parse_date_str(s):
        for fmt in ("%d %b %Y", "%d %B %Y"):
            try:
                return datetime.strptime(s, fmt)
            except ValueError:
                continue
        return None

    wb = load_workbook(DATE_WB_PATH, read_only=True, data_only=True)
    seen_keys = set()
    records = []
    date_set = []
    sector_set = []

    for sname in wb.sheetnames:
        # Validate it's a date sheet
        sheet_date = None
        for fmt in ("%d-%b-%Y", "%d-%B-%Y"):
            try:
                sheet_date = datetime.strptime(sname.strip(), fmt).strftime("%d %b %Y")
                break
            except ValueError:
                continue
        if not sheet_date:
            continue

        ws = wb[sname]
        for row in ws.iter_rows(values_only=True):
            if not row or row[0] is None:
                continue
            sv = str(row[0]).strip()
            if any(sv.startswith(p) for p in SKIP_PREFIXES):
                continue
            if len(row) < 2 or not row[1]:
                continue

            rec_raw = {}
            for ci, col in enumerate(ALL_COLS):
                rec_raw[col] = row[ci] if ci < len(row) else None

            company = str(rec_raw.get("Company Name", "") or "").strip()
            result_date = str(rec_raw.get("Result Date", "") or "").strip()
            quarter = str(rec_raw.get("Quarter", "") or "").strip()
            key = (result_date, company, quarter)
            if not company or key in seen_keys:
                continue
            seen_keys.add(key)

            out = {}
            for col in EXPOSE_COLS:
                v = rec_raw.get(col)
                # Ensure Indices is treated as string even if it's currently None in Excel
                if col == "Indices":
                    out[col] = str(v).strip() if (v is not None and v != "") else ""
                else:
                    out[col] = (str(v).strip() if v is not None else "") if col in STR_COLS else safe_num(v)
            records.append(out)

            if result_date and result_date not in date_set:
                date_set.append(result_date)
            sec = out.get("Sector", "")
            if sec and sec not in sector_set:
                sector_set.append(sec)

    wb.close()

    date_set.sort(key=lambda s: parse_date_str(s) or datetime.min)
    sector_set.sort()

    log(f"[OK] Loaded {len(records)} records across {len(date_set)} dates from Excel")
    return {"records": records, "dates": date_set, "sectors": sector_set}



def _read_sector_dashboard_data():
    """Extract JS data from Sector_Dashboard.html."""
    if not os.path.exists(SECTOR_DASHBOARD):
        return {"sectors": [], "metrics": [], "benchmark": {}, "data": {}}

    with open(SECTOR_DASHBOARD, "r", encoding="utf-8") as f:
        html = f.read()

    import re
    m = re.search(r'const SECTORS\s*=\s*(\[.*?\]);', html)
    sectors = json.loads(m.group(1)) if m else []

    m = re.search(r'const METRICS\s*=\s*(\[.*?\]);', html)
    metrics = json.loads(m.group(1)) if m else []

    m = re.search(r'const PP_METRICS\s*=\s*new Set\((\[.*?\])\)', html)
    pp_metrics = json.loads(m.group(1)) if m else []

    m = re.search(r'const BENCHMARK\s*=\s*(\{.*?\});', html)
    benchmark = json.loads(m.group(1)) if m else {}

    m = re.search(r'const DATA\s*=\s*(\{.*?\});', html, re.DOTALL)
    data = json.loads(m.group(1)) if m else {}

    return {
        "sectors": sectors, "metrics": metrics,
        "pp_metrics": pp_metrics, "benchmark": benchmark, "data": data
    }


def _read_board_meetings_data():
    """Read Board_Meetings.csv and return as list of dicts."""
    if not os.path.exists(BOARD_CSV):
        return []

    import csv
    meetings = []
    with open(BOARD_CSV, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            # Clean up empty trailing columns
            clean = {k.strip(): v.strip() for k, v in row.items()
                     if k and k.strip()}
            if clean.get("Security Code"):
                meetings.append(clean)

    # Parse meeting dates and filter for upcoming (next 45 days)
    today = datetime.today()
    upcoming = []
    for m in meetings:
        md = m.get("Meeting Date", "")
        try:
            dt = None
            for fmt in ("%d %b %Y", "%d/%m/%Y", "%d-%m-%Y"):
                try:
                    dt = datetime.strptime(md.strip(), fmt)
                    break
                except ValueError:
                    continue
            if dt:
                m["_parsed_date"] = dt.strftime("%Y-%m-%d")
                m["_days_away"] = (dt - today).days
                upcoming.append(m)
        except Exception:
            upcoming.append(m)

    # Sort by meeting date
    upcoming.sort(key=lambda x: x.get("_parsed_date", "9999"))
    return upcoming


def _generate_unified_html(daily_data, sector_data, board_data):
    """Generate the unified public dashboard HTML.

    Uses placeholder substitution instead of f-strings to avoid
    {{ / }} escaping issues with CSS and JavaScript curly braces.
    """
    ts = datetime.now().strftime("%d %b %Y, %H:%M")

    # Serialize data for embedding
    daily_json = json.dumps(daily_data["records"])
    daily_dates_json = json.dumps(daily_data["dates"])
    daily_sectors_json = json.dumps(daily_data["sectors"])

    sector_sectors_json = json.dumps(sector_data["sectors"])
    sector_metrics_json = json.dumps(sector_data["metrics"])
    sector_pp_json = json.dumps(sector_data.get("pp_metrics", []))
    sector_benchmark_json = json.dumps(sector_data["benchmark"])
    sector_data_json = json.dumps(sector_data["data"])

    # Clean board data for JSON
    board_clean = []
    for m in board_data:
        board_clean.append({
            "code": m.get("Security Code", ""),
            "name": m.get("Company name", ""),
            "industry": m.get("Industry", ""),
            "purpose": m.get("Purpose", ""),
            "meetingDate": m.get("Meeting Date", ""),
            "announcementDate": m.get("Announcement Date", ""),
            "parsedDate": m.get("_parsed_date", ""),
            "daysAway": m.get("_days_away", 999),
        })
    board_json = json.dumps(board_clean)

    # Build HTML using regular string with placeholder substitution
    # to avoid f-string {{ }} escaping issues with CSS/JS braces
    html = r"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Ashika Results Dashboard</title>
<meta name="description" content="Daily company results, sector analysis, and upcoming board meetings dashboard. MCap ≥ 1000 Cr universe.">
<link href="https://fonts.googleapis.com/css2?family=Outfit:wght@300;400;600;700&family=Inter:wght@400;700&display=swap" rel="stylesheet">
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"></script>
<style>
:root {
  --bg: #0a0e14;
  --card-bg: rgba(23, 28, 36, 0.7);
  --border: rgba(255, 255, 255, 0.1);
  --accent: #3b82f6;
  --accent-purple: #8b5cf6;
  --success: #10b981;
  --error: #ef4444;
  --amber: #f59e0b;
  --cyan: #06b6d4;
  --text: #f1f5f9;
  --text-dim: #94a3b8;
  --glass: blur(12px);
  --shadow: 0 8px 32px rgba(0,0,0,0.4);
}
* { box-sizing: border-box; margin: 0; padding: 0; }
html { scroll-behavior: smooth; }
body { font-family:'Outfit', sans-serif; background: var(--bg); color: var(--text); min-height: 100vh; line-height: 1.6; }
body::before { content:''; position:fixed; top:0;left:0;width:100%;height:100%; background: radial-gradient(ellipse at 20% 50%,rgba(59,130,246,.05) 0%,transparent 50%), radial-gradient(ellipse at 80% 20%,rgba(139,92,246,.04) 0%,transparent 50%); pointer-events:none; z-index:0; }

header {
    background: #0a0e14 !important;
    backdrop-filter: var(--glass);
    border-bottom: 1px solid var(--border);
    padding: 0.75rem 2%;
    position: sticky;
    top: 0;
    z-index: 1000;
    display: grid;
    grid-template-columns: 1fr auto 1fr;
    align-items: center;
}

.header-left { justify-self: start; display: flex; align-items: center; gap: 0.8rem; }
.nav-links { justify-self: center; display: flex; gap: 0.3rem; }
.header-right { justify-self: end; font-size: 0.75rem; color: var(--text-dim); }

.tabs { display:flex; gap:8px; padding:12px 32px; background:rgba(10,14,20,0.8); backdrop-filter:var(--glass); border-bottom:1px solid var(--border); position:sticky; top:54px; z-index:900; }
.tab { padding:10px 24px; border-radius:12px; font-size:.85rem; font-weight:600; cursor:pointer; transition:0.3s; color:var(--text-dim); border:1px solid var(--border); background: var(--card-bg); }
.tab:hover { color:var(--text); background:rgba(255,255,255,0.05); }
.tab.active { background:var(--accent); color:#fff; border-color:var(--accent); box-shadow: 0 0 15px var(--accent); }
.tab .badge-count { display:inline-block; min-width:20px; padding:1px 6px; border-radius:10px; font-size:.7rem; font-weight:700; margin-left:6px; background:rgba(255,255,255,0.2); }

.content { max-width:1400px; margin:0 auto; padding:24px 32px; position:relative; z-index:1; }
.tab-panel { display:none; animation:fadeIn .4s ease; }
.tab-panel.active { display:block; }
@keyframes fadeIn { from{opacity:0;transform:translateY(8px)} to{opacity:1;transform:translateY(0)} }

.card { background:var(--card-bg); border:1px solid var(--border); border-radius:24px; padding:20px; backdrop-filter:var(--glass); margin-bottom:20px; transition:0.3s; }
.card h2 { font-size:1rem; font-weight:700; margin-bottom:16px; color:var(--text); text-transform:uppercase; letter-spacing:0.5px; }

.stats-grid { display:grid; grid-template-columns:repeat(auto-fit,minmax(180px,1fr)); gap:12px; margin-bottom:20px; }
.stat-card { background:var(--card-bg); border:1px solid var(--border); border-radius:16px; padding:16px; text-align:center; transition:0.3s; }
.stat-card:hover { transform:translateY(-2px); border-color: var(--accent); }
.stat-card .value { font-size:1.8rem; font-weight:700; color: var(--text); }
.stat-card .value.green { color: var(--success); }
.stat-card .value.red { color: var(--error); }
.stat-card .label { font-size:.7rem; color:var(--text-dim); margin-top:4px; text-transform:uppercase; font-weight:600; }

.global-filters-bar { background:rgba(10,14,20,0.4); border-bottom:1px solid var(--border); padding:12px 32px; display:flex; flex-wrap:wrap; gap:16px; align-items:center; }
.controls { display:flex; gap:12px; flex-wrap:wrap; align-items:flex-end; margin-bottom:20px; }
.ctrl { display:flex; flex-direction:column; gap:4px; }
.ctrl label { font-size:.65rem; font-weight:700; color:var(--text-dim); text-transform:uppercase; }

select,input[type="text"],input[type="date"],input[type="number"] { padding:10px 14px; border-radius:10px; border:1px solid var(--border); background:#1e293b; color:var(--text); font-family:'Outfit',sans-serif; font-size:.85rem; cursor:pointer; transition:0.3s; color-scheme:dark; }
select:focus,input:focus { outline:none; border-color:var(--accent); box-shadow: 0 0 10px rgba(59,130,246,0.2); }

.btn { padding:10px 18px; border-radius:10px; border:1px solid var(--border); background:var(--card-bg); color:var(--text); font-size:.82rem; font-weight:600; cursor:pointer; transition:0.3s; }
.btn:hover { border-color:var(--accent); color:var(--accent); }
.btn.active { background:var(--accent); color:#fff; border-color:var(--accent); }

.table-wrap { overflow-x:auto; border-radius:18px; border:1px solid var(--border); background: var(--card-bg); }
table { width:100%; border-collapse:collapse; font-size:.85rem; min-width:900px; }
th { background:rgba(255,255,255,0.03); color:var(--text-dim); padding:12px 16px; text-align:center; font-weight:600; font-size:.7rem; text-transform:uppercase; border-bottom:1px solid var(--border); cursor:pointer; position:sticky; top:0; z-index:2; }
th:hover { color: var(--text); background: rgba(255,255,255,0.05); }
td { padding:12px 16px; border-bottom:1px solid rgba(255,255,255,0.03); text-align:center; white-space:nowrap; }
tr:hover td { background:rgba(255,255,255,0.02); }

.badge { display:inline-block; padding:4px 10px; border-radius:8px; font-size:.75rem; font-weight:700; }
.g2 { background:rgba(16,185,129,.1); color:var(--success); border: 1px solid var(--success); }
.g1 { background:rgba(16,185,129,.05); color:var(--success); opacity:0.8; }
.r1 { background:rgba(239,68,68,.05); color:var(--error); opacity:0.8; }
.r2 { background:rgba(239,68,68,.1); color:var(--error); border: 1px solid var(--error); }
.a0 { background:rgba(245,158,11,.1); color:var(--amber); border: 1px solid var(--amber); }

.company-link { color:var(--accent); text-decoration:none; font-weight:600; }
.company-link:hover { text-decoration:underline; }

.stab { padding:8px 16px; border-radius:10px; border:1px solid var(--border); background:var(--card-bg); color:var(--text-dim); font-size:.8rem; font-weight:600; cursor:pointer; transition:0.3s; }
.stab.active { background:var(--accent); color:#fff; border-color:var(--accent); }
input[type="number"] { padding:8px 12px; border-radius:var(--radius-sm); border:1px solid var(--border-glass); background:#1e293b; color:var(--text-primary); font-family:'Inter',sans-serif; font-size:.82rem; cursor:pointer; transition:var(--transition); color-scheme:dark; width:110px; }
input[type="number"]:hover { border-color:var(--accent-blue); }
input[type="number"]:focus { outline:none; border-color:var(--accent-blue); box-shadow:0 0 0 3px rgba(59,130,246,.15); }
.pp-g2 { background:rgba(16,185,129,.2); color:#34d399; }
.pp-g1 { background:rgba(16,185,129,.1); color:#6ee7b7; }
.pp-r1 { background:rgba(239,68,68,.1); color:#fca5a5; }
.pp-r2 { background:rgba(239,68,68,.25); color:#fca5a5; }
.meeting-today { background:rgba(245,158,11,.08)!important; }
.meeting-today td { color:var(--accent-amber)!important; font-weight:600; }
.countdown { display:inline-block; padding:2px 8px; border-radius:4px; font-size:.72rem; font-weight:700; }
.cd-today { background:rgba(245,158,11,.2); color:#fbbf24; }
.cd-soon { background:rgba(59,130,246,.15); color:#93c5fd; }
.cd-later { background:rgba(255,255,255,.05); color:var(--text-muted); }
.cd-past { background:rgba(239,68,68,.1); color:#fca5a5; }
.purpose-tag { display:inline-block; padding:2px 8px; border-radius:4px; font-size:.68rem; font-weight:600; margin:1px; }
.pt-results { background:rgba(16,185,129,.15); color:#6ee7b7; }
.pt-dividend { background:rgba(139,92,246,.15); color:#c4b5fd; }
.pt-general { background:rgba(255,255,255,.05); color:var(--text-muted); }
.pt-other { background:rgba(59,130,246,.1); color:#93c5fd; }
canvas { width:100%!important; }
.chart-wrap { position:relative; height:380px; width:100%; }
#sBarChart { cursor:pointer; }
.footer { text-align:center; padding:20px; font-size:.75rem; color:var(--text-muted); border-top:1px solid var(--border-glass); margin-top:40px; }
.no-data { display:none; padding:60px 20px; text-align:center; color:var(--text-muted); font-size:1rem; }
@media(max-width:900px) {
  table { min-width:600px; font-size:.72rem; }
  th,td { padding:7px 8px; }
}
@media(max-width:768px) {
  .header { padding:14px 16px; }
  .header h1 { font-size:1.1rem; }
  .header .meta { font-size:.72rem; }
  .tabs { padding:8px 12px; gap:4px; overflow-x:auto; flex-wrap:nowrap; -webkit-overflow-scrolling:touch; }
  .tab { padding:8px 14px; font-size:.78rem; white-space:nowrap; flex-shrink:0; }
  .content { padding:12px 14px; }
  .stats-grid { grid-template-columns:repeat(2,1fr); gap:8px; }
  .stat-card { padding:12px 8px; }
  .stat-card .value { font-size:1.2rem; }
  .controls { flex-direction:column; gap:10px; align-items:stretch; }
  .ctrl { width:100%; }
  .ctrl label { font-size:.68rem; }
  select,input[type="text"],input[type="date"] { width:100%!important; font-size:.8rem; box-sizing:border-box; }
  .btn { padding:8px 12px; font-size:.78rem; }
  .ctrl div[style*="flex"] { flex-wrap:wrap; gap:6px; }
  .ctrl div[style*="flex"] .btn { flex:1; min-width:80px; text-align:center; }
  canvas { max-height:280px; min-height:180px; }
  .chart-wrap { height:260px; }
  th,td { padding:6px 7px; }
  canvas { max-height:300px; }
  .card { padding:14px 12px; }
  .footer { padding:14px; font-size:.7rem; }
}
</style>
</head>
<body>
<header>
  <div class="header-left">
      <a href="index.html" class="home-link">🏠</a>
      <div class="logo">Ashika <span>Results</span></div>
  </div>
  <nav class="nav-links">
      <a href="volume_dashboard.html">Quadrants</a>
      <a href="analytics.html">Analytics</a>
      <a href="results_dashboard.html" id="nav-results">Results</a>
      <a href="results_dashboard.html#board" id="nav-board" onclick="switchTab('board')">Board</a>
  </nav>
  <div class="header-right">
      <span class="meta">Updated __TIMESTAMP__</span>
  </div>
</header>
<div class="tabs" id="tabBar">
  <div class="tab active" onclick="switchTab('daily')" id="tab-daily">📅 Daily Results <span class="badge-count" id="dailyCount">0</span></div>
  <div class="tab" onclick="switchTab('sector')" id="tab-sector">📊 Sector Analysis</div>
  <div class="tab" onclick="switchTab('board')" id="tab-board">📋 Board Meetings <span class="badge-count" id="boardCount">0</span></div>
</div>
</div>
<div class="global-filters-bar">
  <div class="ctrl" style="flex-direction:row;align-items:center;gap:12px;">
    <div style="font-weight:700;color:var(--accent-cyan);text-transform:uppercase;font-size:0.75rem;letter-spacing:1px;">Global Filters</div>
    <div style="width:1px;height:24px;background:var(--border-glass)"></div>
  </div>
  <div class="ctrl"><select id="gIndex" onchange="gRefreshAll()"><option value="">All Indices</option><option value="Nifty 50">Nifty 50</option><option value="Nifty 500">Nifty 500</option><option value="Nifty Midcap 150">Nifty Midcap 150</option><option value="Nifty Smallcap 250">Nifty Smallcap 250</option></select></div>
  <div class="ctrl"><div style="display:flex;gap:4px;align-items:center;"><input style="width:110px" type="number" id="gMCap" value="0" min="0" step="100" placeholder="Min MCap (Cr)" oninput="gRefreshAll()"><span style="color:var(--text-muted)">-</span><input style="width:110px" type="number" id="gMCapMax" value="" min="0" step="100" placeholder="Max MCap (Cr)" oninput="gRefreshAll()"></div></div>
</div>
<div class="content">
<!-- Daily Results Panel -->
<div class="tab-panel active" id="panel-daily">
  <div class="stats-grid" id="dailyStats"></div>
  <div class="controls">
    <div class="ctrl"><label>From</label><select id="dDateFrom"></select></div>
    <div class="ctrl"><label>To</label><select id="dDateTo"></select></div>
    <div class="ctrl"><label>&nbsp;</label><button class="btn" onclick="dApplyDate()">Apply</button></div>
    <div class="ctrl"><label>&nbsp;</label><button class="btn" onclick="dResetDate()">All</button></div>
    <div class="ctrl"><label>Sector</label><select id="dSector"><option value="">All Sectors</option></select></div>
    <div class="ctrl"><label>Subsector</label><select id="dSubsector" onchange="dRefresh()"><option value="">All Subsectors</option></select></div>
    <div class="ctrl"><label>Sort By</label><select id="dSort"></select></div>
    <div class="ctrl"><label>Order</label><div style="display:flex;gap:4px"><button class="btn active" id="dBtnDesc" onclick="dSetOrder('desc')">High→Low</button><button class="btn" id="dBtnAsc" onclick="dSetOrder('asc')">Low→High</button></div></div>
    <div class="ctrl"><label>Chart Metric</label><select id="dChartMetric"></select></div>
    <div class="ctrl"><label>Auto Refresh</label><select id="dAutoRefresh" onchange="setAutoRefresh()"><option value="0">Off</option><option value="5">5 min</option><option value="15">15 min</option><option value="30">30 min</option><option value="60">60 min</option></select></div>
    <div class="ctrl"><label>Search</label><input type="text" id="dSearch" placeholder="Company…" oninput="dRefresh()"></div>
  </div>
  <div class="card"><h2 id="dChartTitle">Loading…</h2><div class="chart-wrap"><canvas id="dBarChart"></canvas></div></div>
  <div class="sec-toggle"><button class="stab active" id="stab-fin" onclick="setSec('fin')">Financials</button><button class="stab" id="stab-qoq" onclick="setSec('qoq')">QoQ Growth</button><button class="stab" id="stab-yoy" onclick="setSec('yoy')">YoY Quarterly</button><button class="stab" id="stab-ann" onclick="setSec('ann')">Annual YoY</button></div>
  <div class="card"><div class="no-data" id="dNoData">No companies match the current filters.</div><div class="table-wrap"><table id="dTable"><thead><tr class="grp-hdr" id="dGrpHdr"></tr><tr id="dColHdr"></tr></thead><tbody id="dBody"></tbody></table></div></div>
</div>

<!-- Sector Panel -->
<div class="tab-panel" id="panel-sector">
  <div class="controls">
    <div class="ctrl"><label>Metric</label><select id="sMetric"></select></div>
    <div class="ctrl"><label>View</label><div style="display:flex;gap:4px"><button class="btn active" id="sBtnAbs" onclick="sSetView('absolute')">Absolute</button><button class="btn" id="sBtnDelta" onclick="sSetView('delta')">vs Benchmark Δ</button></div></div>
    <div class="ctrl"><label>Sort</label><div style="display:flex;gap:4px"><button class="btn active" id="sBtnDesc" onclick="sSetSort('desc')">High→Low</button><button class="btn" id="sBtnAsc" onclick="sSetSort('asc')">Low→High</button><button class="btn" id="sBtnAlpha" onclick="sSetSort('alpha')">A–Z</button></div></div>
    <div class="ctrl"><label>Show</label><select id="sTopN"><option value="0">All</option><option value="10">Top 10</option><option value="15">Top 15</option><option value="20">Top 20</option></select></div>
  </div>
  <div class="card"><h2 id="sChartTitle">Loading…</h2><div class="chart-wrap"><canvas id="sBarChart"></canvas></div></div>
  <div class="card"><div class="table-wrap"><table id="sTable"><thead><tr id="sHeadTop"></tr><tr id="sHead"></tr></thead><tbody id="sBody"></tbody></table></div></div>
</div>

<!-- Board Meetings Panel -->
<div class="tab-panel" id="panel-board">
  <div class="stats-grid" id="boardStats"></div>
  <div class="controls">
    <div class="ctrl"><label>Industry</label><select id="bIndustry"><option value="">All Industries</option></select></div>
    <div class="ctrl"><label>Purpose</label><select id="bPurpose"><option value="">All Purposes</option></select></div>
    <div class="ctrl"><label>Time Range</label><select id="bRange" onchange="bClearExact()"><option value="0">Today Only</option><option value="7">Next 7 Days</option><option value="14">Next 14 Days</option><option value="30">Next 30 Days</option><option value="45" selected>Next 45 Days</option><option value="999">All Future</option></select></div>
    <div class="ctrl"><label>Exact Date</label><input type="date" id="bExactDate" oninput="bRefresh()" style="width:150px"></div>
    <div class="ctrl"><label>Search</label><input type="text" id="bSearch" placeholder="Company/Code…" oninput="bRefresh()"></div>
  </div>
  <div class="card"><div class="no-data" id="bNoData">No board meetings match the current filters.</div><div class="table-wrap"><table id="bTable"><thead><tr id="bHead"></tr></thead><tbody id="bBody"></tbody></table></div></div>
</div>
</div>

<div class="footer">All Market Caps · YoY/QoQ in % · Margins in bps (1bps = 0.01pp) · Data from Screener.in & BSE India · Updated __TIMESTAMP__</div>

<script>
const DAILY_DATA=__DAILY_DATA__;
const DAILY_DATES=__DAILY_DATES__;
const DAILY_SECTORS=__DAILY_SECTORS__;
const SEC_SECTORS=__SEC_SECTORS__;
const SEC_METRICS=__SEC_METRICS__;
const SEC_PP=new Set(__SEC_PP__);
let SEC_BENCHMARK={};
let SEC_DATA={};
const BOARD_DATA=__BOARD_DATA__;
// Recalculate daysAway dynamically from parsedDate (fixes timezone/time-of-day offset)
(function(){const t=new Date();t.setHours(0,0,0,0);BOARD_DATA.forEach(m=>{if(m.parsedDate){const d=new Date(m.parsedDate+'T00:00:00');m.daysAway=Math.round((d-t)/86400000);}});})();

function switchTab(n){
  document.querySelectorAll('.tab').forEach(t=>t.classList.remove('active'));
  document.querySelectorAll('.tab-panel').forEach(p=>p.classList.remove('active'));
  document.getElementById('tab-'+n).classList.add('active');
  document.getElementById('panel-'+n).classList.add('active');
  
  // Highlight correct nav link
  document.getElementById('nav-results').classList.toggle('active', n !== 'board');
  document.getElementById('nav-board').classList.toggle('active', n === 'board');
}

window.addEventListener('load', () => {
    const hash = window.location.hash;
    if (hash === '#board') switchTab('board');
    else if (hash === '#sector') switchTab('sector');
    else switchTab('daily');
});

window.addEventListener('hashchange', () => {
    const hash = window.location.hash;
    if (hash === '#board') switchTab('board');
    else if (hash === '#results' || !hash) switchTab('daily');
});

const D_PCT=new Set(["Sales YoY Q%","EBITDA YoY Q%","NP YoY Q%","EPS YoY Q%","Sales QoQ%","EBITDA QoQ%","NP QoQ%","EPS QoQ%","FY Sales YoY%","FY EBITDA YoY%","FY NP YoY%","FY EPS YoY%"]);
const D_PP=new Set(["EBITDA Margin YoY pp","PAT Margin YoY pp","EBITDA Margin QoQ pp","PAT Margin QoQ pp","FY EBITDA Margin YoY pp","FY PAT Margin YoY pp"]);

function badge(v,col){if(v===null||v===undefined)return'<span class="na">—</span>';const n=parseFloat(v);if(isNaN(n))return`<span style="font-size:.72rem">${v}</span>`;const pp=D_PP.has(col),pct=D_PCT.has(col);if(!pp&&!pct)return n.toLocaleString('en-IN',{maximumFractionDigits:0});let c;if(pp){const bps=Math.round(n*100);c=bps>=200?'pp-g2':bps>=0?'pp-g1':bps>=-200?'pp-r1':'pp-r2';return`<span class="badge ${c}">${bps>=0?'+':''}${bps}bps</span>`;}c=n>=15?'g2':n>=5?'g1':n>=0?'a0':n>=-10?'r1':'r2';return`<span class="badge ${c}">${n.toFixed(1)}%</span>`;}

function sBadge(v,isPP){if(v===null||v===undefined)return'<span class="na">—</span>';const n=parseFloat(v);let c;if(isPP){const bps=Math.round(n*100);c=bps>=200?'pp-g2':bps>=0?'pp-g1':bps>=-200?'pp-r1':'pp-r2';return`<span class="badge ${c}">${bps>=0?'+':''}${bps}bps</span>`;}c=n>=15?'g2':n>=5?'g1':n>=0?'a0':n>=-10?'r1':'r2';return`<span class="badge ${c}">${n.toFixed(1)}%</span>`;}

function barColor(v,isPP){if(v===null)return'rgba(100,116,139,.3)';if(isPP){const bps=v*100;return bps>=200?'rgba(16,185,129,.7)':bps>=0?'rgba(16,185,129,.4)':bps>=-200?'rgba(239,68,68,.4)':'rgba(239,68,68,.7)';}return v>=15?'rgba(16,185,129,.7)':v>=5?'rgba(16,185,129,.4)':v>=0?'rgba(245,158,11,.4)':v>=-10?'rgba(239,68,68,.4)':'rgba(239,68,68,.7)';}

// ── DAILY RESULTS ──
const D_COLS=["Result Date","Company Name","Sector","Indices","Market Cap (Cr)","Quarter","Sales (Cr)","EBITDA (Cr)","Net Profit (Cr)","EPS (Rs)","EBITDA Margin%","PAT Margin%","Sales YoY Q%","EBITDA YoY Q%","NP YoY Q%","EPS YoY Q%","EBITDA Margin YoY pp","PAT Margin YoY pp","Sales QoQ%","EBITDA QoQ%","NP QoQ%","EPS QoQ%","EBITDA Margin QoQ pp","PAT Margin QoQ pp","FY Sales YoY%","FY EBITDA YoY%","FY NP YoY%","FY EPS YoY%","FY EBITDA Margin YoY pp","FY PAT Margin YoY pp"];
const D_SORT=["Sales YoY Q%","EBITDA YoY Q%","NP YoY Q%","EPS YoY Q%","EBITDA Margin YoY pp","PAT Margin YoY pp","Sales QoQ%","NP QoQ%","EBITDA Margin QoQ pp","PAT Margin QoQ pp","FY Sales YoY%","FY NP YoY%","Market Cap (Cr)","Company Name"];
const D_GROUPS=[{label:"Identity",cols:["Result Date","Company Name","Sector","Indices","Market Cap (Cr)","Quarter"]},{label:"Quarterly Financials",cols:["Sales (Cr)","EBITDA (Cr)","Net Profit (Cr)","EPS (Rs)","EBITDA Margin%","PAT Margin%"]},{label:"YoY Growth",cols:["Sales YoY Q%","EBITDA YoY Q%","NP YoY Q%","EPS YoY Q%","EBITDA Margin YoY pp","PAT Margin YoY pp"]},{label:"QoQ Growth",cols:["Sales QoQ%","EBITDA QoQ%","NP QoQ%","EPS QoQ%","EBITDA Margin QoQ pp","PAT Margin QoQ pp"]},{label:"Annual FY YoY",cols:["FY Sales YoY%","FY EBITDA YoY%","FY NP YoY%","FY EPS YoY%","FY EBITDA Margin YoY pp","FY PAT Margin YoY pp"]}];
const D_CHART_M=D_SORT.filter(m=>D_COLS.includes(m)&&m!=="Company Name");
let dSort='Sales YoY Q%',dOrder='desc',dFrom=null,dTo=null,dChart=null;
let dSec='fin';
const D_SEC_GROUPS={
  fin:['Identity','Quarterly Financials'],
  qoq:['Identity','QoQ Growth'],
  yoy:['Identity','YoY Growth'],
  ann:['Identity','Annual FY YoY']
};
function setSec(s){dSec=s;document.querySelectorAll('.stab').forEach(b=>b.classList.remove('active'));document.getElementById('stab-'+s).classList.add('active');dBuildHeaders();dRenderTable(dGetFiltered());}
function dVisibleGroups(){return D_GROUPS.filter(g=>D_SEC_GROUPS[dSec].includes(g.label));}
let autoRefTimer=null;
function setAutoRefresh(){if(autoRefTimer)clearInterval(autoRefTimer);const m=parseInt(document.getElementById('dAutoRefresh').value);if(m>0)autoRefTimer=setInterval(()=>location.reload(),m*60000);}
function dRefreshSubsectors(){
  const sec=document.getElementById('dSector').value;
  const sub=document.getElementById('dSubsector');
  const prev=sub.value;
  // Collect subsectors from data matching sector filter
  const subs=new Set();
  DAILY_DATA.forEach(r=>{if((!sec||r.Sector===sec)&&r.Subsector&&r.Subsector!=='')subs.add(r.Subsector);});
  sub.innerHTML='<option value="">All Subsectors</option>';
  [...subs].sort().forEach(s=>sub.innerHTML+=`<option value="${s}">${s}</option>`);
  if([...subs].includes(prev))sub.value=prev;
}

(function(){
  const df=document.getElementById('dDateFrom'),dt=document.getElementById('dDateTo');
  DAILY_DATES.forEach(d=>{df.innerHTML+=`<option value="${d}">${d}</option>`;dt.innerHTML+=`<option value="${d}">${d}</option>`;});
  if(DAILY_DATES.length){df.value=DAILY_DATES[0];dt.value=DAILY_DATES[DAILY_DATES.length-1];dFrom=DAILY_DATES[0];dTo=DAILY_DATES[DAILY_DATES.length-1];}
  const ss=document.getElementById('dSector');DAILY_SECTORS.forEach(s=>ss.innerHTML+=`<option value="${s}">${s}</option>`);ss.onchange=()=>{dRefreshSubsectors();dRefresh();};
  dRefreshSubsectors();
  const so=document.getElementById('dSort');D_SORT.filter(c=>D_COLS.includes(c)).forEach(c=>so.innerHTML+=`<option value="${c}">${c}</option>`);so.onchange=e=>{dSort=e.target.value;const _cm=document.getElementById('dChartMetric');if(D_CHART_M.includes(dSort))_cm.value=dSort;dRefresh();};
  const cm=document.getElementById('dChartMetric');D_CHART_M.forEach(m=>cm.innerHTML+=`<option value="${m}">${m}</option>`);cm.onchange=()=>dRefresh();
  document.getElementById('dailyCount').textContent=DAILY_DATA.length;
})();

function dApplyDate(){dFrom=document.getElementById('dDateFrom').value;dTo=document.getElementById('dDateTo').value;dRefresh();}
function dResetDate(){dFrom=DAILY_DATES[0]||null;dTo=DAILY_DATES[DAILY_DATES.length-1]||null;if(DAILY_DATES.length){document.getElementById('dDateFrom').value=DAILY_DATES[0];document.getElementById('dDateTo').value=DAILY_DATES[DAILY_DATES.length-1];}dRefresh();}
function dSetOrder(o){dOrder=o;document.getElementById('dBtnDesc').className=o==='desc'?'btn active':'btn';document.getElementById('dBtnAsc').className=o==='asc'?'btn active':'btn';dRefresh();}

function gGetGlobalFiltered(){
  const mFloor=parseFloat(document.getElementById('gMCap').value)||0;
  const mCeil=parseFloat(document.getElementById('gMCapMax').value);
  const idx=document.getElementById('gIndex').value;
  return DAILY_DATA.filter(r=>{
    if(mFloor>0&&(r["Market Cap (Cr)"]==null||parseFloat(r["Market Cap (Cr)"])<mFloor))return false;
    if(!isNaN(mCeil)&&(r["Market Cap (Cr)"]==null||parseFloat(r["Market Cap (Cr)"])>mCeil))return false;
    if(idx && !(r["Indices"]||"").split(',').map(s=>s.trim()).includes(idx))return false;
    return true;
  });
}

function gRefreshAll(){
  const rows = gGetGlobalFiltered();
  SEC_DATA = {};
  SEC_BENCHMARK = {};
  SEC_METRICS.forEach(m => {
    SEC_DATA[m] = [];
    const rawColName = m.replace(/^Avg /, '');
    const benchVals = rows.map(r => parseFloat(r[rawColName])).filter(v => !isNaN(v));
    SEC_BENCHMARK[m] = benchVals.length ? benchVals.reduce((a,b)=>a+b,0)/benchVals.length : null;
    SEC_SECTORS.forEach(sec => {
      const secRows = rows.filter(r => r.Sector === sec);
      const secVals = secRows.map(r => parseFloat(r[rawColName])).filter(v => !isNaN(v));
      SEC_DATA[m].push(secVals.length ? secVals.reduce((a,b)=>a+b,0)/secVals.length : null);
    });
  });

  if(typeof dRefresh === 'function') dRefresh();
  if(typeof sRefresh === 'function') sRefresh();
  
  const dBadge = document.getElementById('dailyCount');
  if(dBadge) dBadge.textContent = rows.length;
}

function dGetFiltered(){
  const sec=document.getElementById('dSector').value;const q=document.getElementById('dSearch').value.toLowerCase();
  const subF=document.getElementById('dSubsector').value;
  let rows=gGetGlobalFiltered().filter(r=>{
    if(dFrom&&dTo){const di=DAILY_DATES.indexOf(r["Result Date"]);if(di<DAILY_DATES.indexOf(dFrom)||di>DAILY_DATES.indexOf(dTo))return false;}
    if(sec&&r.Sector!==sec)return false;
    if(q&&!(r["Company Name"]||'').toLowerCase().includes(q))return false;
    if(subF&&r.Subsector!==subF)return false;
    return true;
  });
  rows.sort((a,b)=>{const av=a[dSort],bv=b[dSort];if(av===null&&bv===null)return 0;if(av===null)return 1;if(bv===null)return -1;if(typeof av==='string')return dOrder==='asc'?av.localeCompare(bv):bv.localeCompare(av);return dOrder==='asc'?av-bv:bv-av;});return rows;
}

function dRenderStats(rows){
  const n=rows.length;
  const secs=new Set(rows.map(r=>r.Sector).filter(Boolean)).size;
  const metric=dSort;
  const isPP=D_PP.has(metric);
  const lbl=metric.replace(' YoY Q%',' YoY').replace(' QoQ%',' QoQ').replace(' YoY pp',' YoY').replace(' QoQ pp',' QoQ').replace('EBITDA Margin','EM').replace('PAT Margin','PM').replace('FY ','').replace(' YoY%',' YoY').replace('Market Cap (Cr)','MCap').replace('Company Name','Name');
  const vld=rows.filter(r=>r[metric]!==null&&r[metric]!==undefined&&!isNaN(parseFloat(r[metric])));
  const pos=vld.filter(r=>parseFloat(r[metric])>=0).length;
  const neg=vld.filter(r=>parseFloat(r[metric])<0).length;
  const avgMv=vld.length?vld.reduce((s,r)=>s+parseFloat(r[metric]),0)/vld.length:null;
  const vldNP=rows.filter(r=>r['NP YoY Q%']!==null&&!isNaN(parseFloat(r['NP YoY Q%'])));
  const avgNPv=vldNP.length?vldNP.reduce((s,r)=>s+parseFloat(r['NP YoY Q%']),0)/vldNP.length:null;
  function fmt(v,pp){if(v===null)return'—';if(pp)return(Math.round(v*100)>=0?'+':'')+Math.round(v*100)+'bps';return(v>=0?'+':'')+v.toFixed(1)+'%';}
  document.getElementById('dailyStats').innerHTML=
    `<div class="stat-card"><div class="value">${n}</div><div class="label">Companies</div></div>`+
    `<div class="stat-card"><div class="value">${secs}</div><div class="label">Sectors</div></div>`+
    `<div class="stat-card"><div class="value green">${pos}</div><div class="label">${lbl} ≥ 0</div></div>`+
    `<div class="stat-card"><div class="value red">${neg}</div><div class="label">${lbl} < 0</div></div>`+
    `<div class="stat-card"><div class="value">${fmt(avgMv,isPP)}</div><div class="label">Avg ${lbl}</div></div>`+
    `<div class="stat-card"><div class="value">${fmt(avgNPv,false)}</div><div class="label">Avg NP YoY</div></div>`;
}

function dRenderChart(rows){
  const metric=document.getElementById('dChartMetric').value||D_CHART_M[0];if(!metric)return;
  const secMap={};rows.forEach(r=>{const s=r.Sector||'Unknown';if(!secMap[s])secMap[s]={vals:[],n:0};if(r[metric]!==null&&r[metric]!==undefined){secMap[s].vals.push(r[metric]);}secMap[s].n++;});
  let pairs=Object.entries(secMap).map(([s,d])=>({s,v:d.vals.length?d.vals.reduce((a,b)=>a+b,0)/d.vals.length:null,n:d.n,nv:d.vals.length})).filter(p=>p.v!==null).sort((a,b)=>b.v-a.v);
  document.getElementById('dChartTitle').textContent=`${metric} — Avg by Sector (${rows.length} companies)`;
  if(dChart)dChart.destroy();const isPP=D_PP.has(metric);
  dChart=new Chart(document.getElementById('dBarChart'),{type:'bar',data:{labels:pairs.map(p=>p.s),datasets:[{label:metric,data:pairs.map(p=>isPP?Math.round(p.v*100):parseFloat(p.v.toFixed(1))),backgroundColor:pairs.map(p=>barColor(p.v,isPP)),borderRadius:6,borderSkipped:false}]},options:{indexAxis:'y',maintainAspectRatio:false,responsive:true,plugins:{legend:{display:false},tooltip:{backgroundColor:'rgba(17,24,39,.95)',titleColor:'#f1f5f9',bodyColor:'#94a3b8',borderColor:'rgba(59,130,246,.3)',borderWidth:1,callbacks:{label:ctx=>{const p=pairs[ctx.dataIndex];return` ${ctx.parsed.x!==null?ctx.parsed.x+(isPP?'bps':'%'):'—'} (${p.nv}/${p.n} cos.)`;}}}},onClick:(evt,items)=>{if(items.length>0){const sec=pairs[items[0].index].s;const sel=document.getElementById('dSector');sel.value=sec;dRefresh();}},scales:{x:{title:{display:true,text:isPP?'bps':'%',color:'#64748b'},grid:{color:'rgba(255,255,255,.04)'},ticks:{color:'#94a3b8'}},y:{ticks:{font:{size:10},color:'#94a3b8'},grid:{display:false}}}}});
}

function dBuildHeaders(){
  const exist=D_COLS;const g=document.getElementById('dGrpHdr'),c=document.getElementById('dColHdr');let t='',s='';
  dVisibleGroups().forEach(gr=>{const vis=gr.cols.filter(c=>exist.includes(c));if(!vis.length)return;t+=`<th colspan="${vis.length}">${gr.label}</th>`;vis.forEach(col=>{const sh=col.replace(' YoY Q%','').replace(' QoQ%','').replace(' YoY%','').replace(' YoY pp',' bps').replace(' QoQ pp',' bps').replace(' (Cr)','').replace(' (Rs)','').replace('Net Profit','NP').replace('Market Cap','MCap').replace('Result ','');s+=`<th onclick="dTableSort('${col}')" id="dth_${col.replace(/[^a-zA-Z0-9]/g,'_')}">${sh}</th>`;});});
  g.innerHTML=t;c.innerHTML=s;
}
let dTSortCol=null,dTSortDir='desc';
function dTableSort(col){if(dTSortCol===col)dTSortDir=dTSortDir==='desc'?'asc':'desc';else{dTSortCol=col;dTSortDir='desc';}document.querySelectorAll('#dColHdr th').forEach(t=>t.classList.remove('sort-asc','sort-desc'));const el=document.getElementById('dth_'+col.replace(/[^a-zA-Z0-9]/g,'_'));if(el)el.classList.add(dTSortDir==='desc'?'sort-desc':'sort-asc');dRenderTable(dGetFiltered());}

function dRenderTable(rows){
  if(dTSortCol){rows=[...rows].sort((a,b)=>{const av=a[dTSortCol],bv=b[dTSortCol];if(av===null&&bv===null)return 0;if(av===null)return 1;if(bv===null)return -1;if(typeof av==='string')return dTSortDir==='asc'?av.localeCompare(bv):bv.localeCompare(av);return dTSortDir==='asc'?av-bv:bv-av;});}
  const nd=document.getElementById('dNoData'),tb=document.getElementById('dTable');if(!rows.length){nd.style.display='block';tb.style.display='none';return;}nd.style.display='none';tb.style.display='';
  let h='';rows.forEach(r=>{h+='<tr>';dVisibleGroups().forEach(g=>g.cols.filter(c=>D_COLS.includes(c)).forEach(c=>{if(c==='Company Name'){const tk=(r['Screener Ticker']||'').replace(/^None$/i,'');const url=tk?`https://www.screener.in/company/${tk}/`:'';h+=url?`<td><a class="company-link" href="${url}" target="_blank" rel="noopener">${r[c]||''}</a></td>`:`<td>${r[c]||''}</td>`;}else{h+=`<td>${badge(r[c],c)}</td>`;};}));h+='</tr>';});
  document.getElementById('dBody').innerHTML=h;
}
dBuildHeaders();

// ── SECTOR TAB ──
const S_COL_GROUPS=[{label:"Qtr Growth (YoY)",cols:["Avg Sales YoY Q%","Avg EBITDA YoY Q%","Avg NP YoY Q%","Avg EPS YoY Q%"]},{label:"Qtr Margins",cols:["Avg EBITDA Margin%","Avg PAT Margin%"]},{label:"Qtr Margin Δ",cols:["Avg EBITDA Margin YoY pp","Avg PAT Margin YoY pp","Avg EBITDA Margin QoQ pp","Avg PAT Margin QoQ pp"]},{label:"Annual Growth",cols:["Avg FY Sales YoY%","Avg FY EBITDA YoY%","Avg FY NP YoY%","Avg FY EPS YoY%"]},{label:"Annual Margins",cols:["Avg FY EBITDA Margin%","Avg FY PAT Margin%","Avg FY EBITDA Margin YoY pp","Avg FY PAT Margin YoY pp"]}];
let sMetric=SEC_METRICS[0]||'',sView='absolute',sSort='desc',sTopN=0,sChart=null;
function sVal(sec,m){const i=SEC_SECTORS.indexOf(sec);const a=SEC_DATA[m];return(a&&i>=0)?a[i]:null;}

(function(){
  const sel=document.getElementById('sMetric');
  const groups=[{l:'Qtr Growth',m:SEC_METRICS.slice(0,4)},{l:'Qtr Margins',m:SEC_METRICS.slice(4,6)},{l:'Qtr Margin Δ',m:SEC_METRICS.slice(6,10)},{l:'Annual Growth',m:SEC_METRICS.slice(10,14)},{l:'Annual Margins',m:SEC_METRICS.slice(14)}];
  groups.forEach(g=>{const og=document.createElement('optgroup');og.label=g.l;g.m.forEach(m=>{const o=document.createElement('option');o.value=m;o.textContent=m;og.appendChild(o);});sel.appendChild(og);});
  sel.onchange=e=>{sMetric=e.target.value;sRefresh();};
  document.getElementById('sTopN').onchange=e=>{sTopN=parseInt(e.target.value);sRefresh();};
})();

function sSetView(v){sView=v;document.getElementById('sBtnAbs').className=v==='absolute'?'btn active':'btn';document.getElementById('sBtnDelta').className=v==='delta'?'btn active':'btn';sRefresh();}
function sSetSort(s){sSort=s;['sBtnDesc','sBtnAsc','sBtnAlpha'].forEach(id=>document.getElementById(id).className='btn');document.getElementById(s==='desc'?'sBtnDesc':s==='asc'?'sBtnAsc':'sBtnAlpha').className='btn active';sRefresh();}

function sGetSorted(){
  let pairs=SEC_SECTORS.map(s=>({sector:s,value:sView==='delta'?(sVal(s,sMetric)!==null&&SEC_BENCHMARK[sMetric]!==null?parseFloat((sVal(s,sMetric)-SEC_BENCHMARK[sMetric]).toFixed(1)):null):sVal(s,sMetric)}));
  pairs = pairs.filter(p=>p.value!==null);
  if(sSort==='desc')pairs.sort((a,b)=>{if(a.value===null)return 1;if(b.value===null)return -1;return b.value-a.value;});
  else if(sSort==='asc')pairs.sort((a,b)=>{if(a.value===null)return 1;if(b.value===null)return -1;return a.value-b.value;});
  else pairs.sort((a,b)=>a.sector.localeCompare(b.sector));
  if(sTopN>0)pairs=pairs.slice(0,sTopN);return pairs;
}

function sRenderChart(){
  const pairs=sGetSorted(),labels=pairs.map(p=>p.sector),values=pairs.map(p=>p.value);
  const bench=SEC_BENCHMARK[sMetric],isPP=SEC_PP.has(sMetric),unit=isPP?'bps':'%';
  const ds=[{label:sView==='delta'?'Delta':isPP?'Avg bps':'Avg %',data:values.map(v=>v===null?null:(isPP?Math.round(v*100):v)),backgroundColor:values.map(v=>barColor(v,isPP)),borderRadius:6,borderSkipped:false}];
  if(sView==='absolute'&&bench!==null)ds.push({type:'line',label:'Market Avg',data:Array(labels.length).fill(isPP?Math.round(bench*100):bench),borderColor:'rgba(245,158,11,.8)',borderWidth:2,borderDash:[6,3],pointRadius:0,fill:false});
  document.getElementById('sChartTitle').textContent=sView==='delta'?`${sMetric} — Delta vs Market (${bench!==null?(isPP?Math.round(bench*100)+'bps':bench.toFixed(1)+'%'):'N/A'})`:`${sMetric} — Sector Averages`;
  if(sChart)sChart.destroy();
  sChart=new Chart(document.getElementById('sBarChart'),{type:'bar',data:{labels,datasets:ds},options:{indexAxis:'y',maintainAspectRatio:false,responsive:true,plugins:{legend:{display:true,position:'top',labels:{color:'#94a3b8'}},tooltip:{backgroundColor:'rgba(17,24,39,.95)',titleColor:'#f1f5f9',bodyColor:'#94a3b8',borderColor:'rgba(59,130,246,.3)',borderWidth:1,callbacks:{label:ctx=>' '+(ctx.parsed.x!==null?ctx.parsed.x+(isPP?'bps':'%'):'—')}}},onClick:(evt,items)=>{if(items.length>0){const sec=labels[items[0].index];switchTab('daily');const sel=document.getElementById('dSector');if(sel)sel.value=sec;dRefresh();}},scales:{x:{title:{display:true,text:isPP?'bps':'%',color:'#64748b'},grid:{color:'rgba(255,255,255,.04)'},ticks:{color:'#94a3b8'}},y:{ticks:{font:{size:10},color:'#94a3b8'},grid:{display:false}}}}});
}

function sRenderTable(){
  const bench=SEC_BENCHMARK[sMetric];const ht=document.getElementById('sHeadTop'),hd=document.getElementById('sHead');
  let top='<th rowspan="2" style="text-align:left">Sector</th>';S_COL_GROUPS.forEach(g=>top+=`<th colspan="${g.cols.length}" style="border-bottom:1px solid rgba(255,255,255,.1)">${g.label}</th>`);top+='<th rowspan="2">vs Bench</th>';ht.innerHTML=top;
  let sub='';S_COL_GROUPS.forEach(g=>g.cols.forEach(c=>sub+=`<th>${c.replace('Avg ','').replace(' Q%','%')}</th>`));hd.innerHTML=sub;
  let brow='<tr style="background:rgba(59,130,246,.1)"><td style="text-align:left;font-weight:700;color:var(--accent-cyan)">▶ Market Avg</td>';
  SEC_METRICS.forEach(m=>{const bv=SEC_BENCHMARK[m];const pp=SEC_PP.has(m);brow+=`<td style="font-weight:700;color:var(--accent-cyan)">${bv!==null?(pp?`${Math.round(bv*100)>=0?'+':''}${Math.round(bv*100)}bps`:`${(bv>=0?'':'')}${bv.toFixed(1)}%`):'—'}</td>`;});brow+='<td>—</td></tr>';
  const pairs=sGetSorted();let rows='';
  pairs.forEach(p=>{const s=p.sector;rows+=`<tr><td style="text-align:left;font-weight:600">${s}</td>`;SEC_METRICS.forEach(m=>rows+=`<td>${sBadge(sVal(s,m),SEC_PP.has(m))}</td>`);const dv=(sVal(s,sMetric)!==null&&bench!==null)?sVal(s,sMetric)-bench:null;const dvDisp=dv!==null?(SEC_PP.has(sMetric)?Math.round(dv*100)+'bps':dv.toFixed(1)+'%'):null;rows+=`<td>${dvDisp!==null?`<span style="color:${dv>=0?'var(--accent-green)':'var(--accent-red)'};font-weight:700">${dv>=0?'+':''}${dvDisp}</span>`:'<span class="na">—</span>'}</td></tr>`;});
  document.getElementById('sBody').innerHTML=brow+rows;
}
function sRefresh(){sRenderChart();sRenderTable();}
function dRefresh(){const r=dGetFiltered();dRenderStats(r);dRenderChart(r);dRenderTable(r);}

// ── BOARD MEETINGS ──
(function(){
  const industries=new Set(BOARD_DATA.map(m=>m.industry).filter(Boolean));const iSel=document.getElementById('bIndustry');
  [...industries].sort().forEach(i=>iSel.innerHTML+=`<option value="${i}">${i}</option>`);iSel.onchange=()=>bRefresh();
  const purposes=new Set();BOARD_DATA.forEach(m=>{(m.purpose||'').split(';').forEach(p=>{const t=p.trim();if(t)purposes.add(t);});});
  const pSel=document.getElementById('bPurpose');[...purposes].sort().forEach(p=>pSel.innerHTML+=`<option value="${p}">${p}</option>`);pSel.onchange=()=>bRefresh();
  document.getElementById('bRange').onchange=()=>bRefresh();
  const futureCount=BOARD_DATA.filter(m=>m.daysAway>=0).length;
  document.getElementById('boardCount').textContent=futureCount;
})();

function purposeTags(p){if(!p)return'';return p.split(';').map(s=>{const t=s.trim();if(!t)return'';const low=t.toLowerCase();let cls='pt-other';if(low.includes('result')||low.includes('audited'))cls='pt-results';else if(low.includes('dividend'))cls='pt-dividend';else if(low.includes('general'))cls='pt-general';return`<span class="purpose-tag ${cls}">${t}</span>`;}).join(' ');}
function countdownBadge(d){if(d===undefined||d===null||d===999)return'<span class="na">—</span>';if(d===0)return'<span class="countdown cd-today">TODAY</span>';if(d<0)return`<span class="countdown cd-past">${d}d</span>`;if(d<=7)return`<span class="countdown cd-soon">${d}d</span>`;return`<span class="countdown cd-later">${d}d</span>`;}

function bClearExact(){document.getElementById('bExactDate').value='';bRefresh();}
function bGetFiltered(){
  const ind=document.getElementById('bIndustry').value;const pur=document.getElementById('bPurpose').value;
  const range=parseInt(document.getElementById('bRange').value);const exactDate=document.getElementById('bExactDate').value;
  const q=document.getElementById('bSearch').value.toLowerCase();
  return BOARD_DATA.filter(m=>{
    if(m.daysAway<0)return false; // always hide past meetings
    if(ind&&m.industry!==ind)return false;
    if(pur&&!(m.purpose||'').includes(pur))return false;
    if(exactDate){if(m.parsedDate!==exactDate)return false;} // exact date overrides range
    else if(range<999&&m.daysAway>range)return false;
    if(q&&!(m.name||'').toLowerCase().includes(q)&&!(m.code||'').toLowerCase().includes(q))return false;
    return true;
  });
}

function bRenderStats(rows){
  const total=rows.length;const today=rows.filter(r=>r.daysAway===0).length;const week=rows.filter(r=>r.daysAway>=0&&r.daysAway<=7).length;
  const results=rows.filter(r=>(r.purpose||'').toLowerCase().includes('result')||(r.purpose||'').toLowerCase().includes('audited')).length;
  const industries=new Set(rows.map(r=>r.industry).filter(Boolean)).size;
  document.getElementById('boardStats').innerHTML=`<div class="stat-card"><div class="value">${total}</div><div class="label">Total Meetings</div></div><div class="stat-card"><div class="value" style="${today>0?'background:linear-gradient(135deg,#f59e0b,#fbbf24);-webkit-background-clip:text;background-clip:text':''}">${today}</div><div class="label">Today</div></div><div class="stat-card"><div class="value">${week}</div><div class="label">This Week</div></div><div class="stat-card"><div class="value green">${results}</div><div class="label">Results Announcements</div></div><div class="stat-card"><div class="value">${industries}</div><div class="label">Industries</div></div>`;
}

function bRenderTable(rows){
  const nd=document.getElementById('bNoData'),tb=document.getElementById('bTable');if(!rows.length){nd.style.display='block';tb.style.display='none';return;}nd.style.display='none';tb.style.display='';
  document.getElementById('bHead').innerHTML='<th style="text-align:left">Company</th><th>Code</th><th style="text-align:left">Industry</th><th>Meeting Date</th><th>Countdown</th><th style="text-align:left">Purpose</th><th>Announced</th>';
  let h='';rows.forEach(r=>{const isToday=r.daysAway===0;h+=`<tr class="${isToday?'meeting-today':''}"><td style="text-align:left;font-weight:600">${r.name}</td><td style="font-size:.72rem;color:var(--text-muted)">${r.code}</td><td style="text-align:left;font-size:.72rem">${r.industry}</td><td style="font-weight:600">${r.meetingDate}</td><td>${countdownBadge(r.daysAway)}</td><td style="text-align:left;white-space:normal;max-width:300px">${purposeTags(r.purpose)}</td><td style="font-size:.72rem;color:var(--text-muted)">${r.announcementDate}</td></tr>`;});
  document.getElementById('bBody').innerHTML=h;
}
function bRefresh(){const r=bGetFiltered();bRenderStats(r);bRenderTable(r);}
bRefresh();

// Trigger global recalculation
gRefreshAll();
</script>
</body>
</html>"""

    # Substitute placeholders with actual data
    html = html.replace("__TIMESTAMP__", ts)
    html = html.replace("__DAILY_DATA__", daily_json)
    html = html.replace("__DAILY_DATES__", daily_dates_json)
    html = html.replace("__DAILY_SECTORS__", daily_sectors_json)
    html = html.replace("__SEC_SECTORS__", sector_sectors_json)
    html = html.replace("__SEC_METRICS__", sector_metrics_json)
    html = html.replace("__SEC_PP__", sector_pp_json)
    html = html.replace("__SEC_BENCHMARK__", sector_benchmark_json)
    html = html.replace("__SEC_DATA__", sector_data_json)
    html = html.replace("__BOARD_DATA__", board_json)

    out_path = os.path.join(PUBLIC_SITE_DIR, "results_dashboard.html")
    with open(out_path, "w", encoding="utf-8") as f:
        f.write(html)
    log(f"[OK] Written: {out_path}")


def git_push():
    """Step 5: Git commit & push to GitHub Pages."""
    log("=" * 50)
    log("STEP 5: Git Push (if configured)")
    log("=" * 50)

    if not os.path.exists(os.path.join(PUBLIC_SITE_DIR, ".git")):
        log("Git not initialized in public_site/ — skipping push.")
        log("Run 'git init' in public_site/ and add a GitHub remote to enable auto-push.")
        return False

    try:
        subprocess.run(["git", "add", "."], cwd=PUBLIC_SITE_DIR,
                       capture_output=True, timeout=30)
        ts = datetime.now().strftime("%Y-%m-%d %H:%M")
        subprocess.run(["git", "commit", "-m", f"Auto-update {ts}"],
                       cwd=PUBLIC_SITE_DIR, capture_output=True, timeout=30)
        result = subprocess.run(["git", "push"], cwd=PUBLIC_SITE_DIR,
                                capture_output=True, text=True, timeout=60)
        if result.returncode == 0:
            log("[OK] Pushed to GitHub Pages")
            log("[OK] Live at: https://arhamsaraogi-star.github.io/results-dashboard/")
            return True
        else:
            log(f"Git push failed: {result.stderr}", "WARN")
            return False
    except Exception as e:
        log(f"Git push error: {e}", "WARN")
        return False


def save_run_log():
    """Save a log of this run."""
    data = {
        "last_run": datetime.now().isoformat(),
        "python": PYTHON,
    }
    with open(RUN_LOG, "w") as f:
        json.dump(data, f, indent=2)


def main():
    print("\n" + "=" * 60)
    print("  DAILY RESULTS PIPELINE")
    print(f"  {datetime.now().strftime('%d %b %Y, %H:%M')}")
    print("=" * 60)

    skip_bse = "--skip-bse" in sys.argv
    force_all = "--force-all" in sys.argv

    # Step 1: BSE Board Meetings & Indices
    if not skip_bse:
        run_indices_scraper() # Refresh index constituents mapping first
        run_bse_scraper()
        # Don't run board.py immediately — it's slow and we can use the CSV directly
        # run_board_processor()

    # Step 2: Determine dates to process
    if force_all:
        dates = []
        d = SEASON_START
        yesterday = datetime.today() - timedelta(days=1)
        while d <= yesterday:
            dates.append(d)
            d += timedelta(days=1)
        log(f"Force-all: {len(dates)} dates from {SEASON_START.strftime('%d %b')}")
    else:
        dates = get_dates_to_process()

    # Step 3: Run sorter for missed dates
    if dates:
        run_sorter(dates)
    else:
        log("No missed dates to process.")

    # Step 4: Build public website
    build_public_site()

    # Step 5: Git push
    git_push()

    # Save run log
    save_run_log()

    print("\n" + "=" * 60)
    print("  PIPELINE COMPLETE")
    print("=" * 60)


if __name__ == "__main__":
    main()
