"""
Board Meetings Workflow
-----------------------
Place Board_Meetings.csv and 6_0_BSE_1000_Sector_Allocation___Results_Schedule.xlsx
in the same folder, then run:
    py board.py

Steps:
  1. Reads Board_Meetings.csv
  2. Filters results announcements
  3. Fetches Market Caps from Screener (login → search BSE code → get ticker → scrape MCap)
  4. Applies 1000 Cr MCap floor filter
  5. Attaches Industry New Name / Igroup / ISubgroup from 6.0 classification file
  6. Writes Board_Meetings_Report.xlsx with 3 sheets
"""

import requests
import re
import time
import os
import sys
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DIR = os.path.dirname(os.path.abspath(__file__))
CSV_IN = os.path.join(DIR, "Board_Meetings.csv")
INDUSTRY_FILE = os.path.join(
    DIR, "6_0_BSE_1000_Sector_Allocation___Results_Schedule.xlsx")
EXCEL_OUT = os.path.join(DIR, "Board_Meetings_Report.xlsx")

SCREENER_EMAIL = "asutosh@ashikagroup.com"
SCREENER_PASSWORD = "Dilipsir@1234"
BASE_URL = "https://www.screener.in"

NAVY = "0E2841"
BLUE = "243F8D"
WHITE = "FFFFFF"
LGRAY = "F2F2F2"
MGRAY = "D9D9D9"
MCAP_FLOOR = 1000  # Cr

RESULTS_KEYWORDS = ["quarterly results", "audited results", "half yearly results",
                    "unaudited results", "annual results"]


# ══════════════════════════════════════════════════════════════════════════
# INDUSTRY CLASSIFICATION
# ══════════════════════════════════════════════════════════════════════════

def load_industry_map():
    """Returns dict {bse_code_str: {'Industry New Name':..., 'Igroup Name':..., 'ISubgroup Name':...}}"""
    if not os.path.exists(INDUSTRY_FILE):
        print(f"  WARNING: Industry file not found: {INDUSTRY_FILE}")
        return {}
    df = pd.read_excel(INDUSTRY_FILE, sheet_name="Industry - BSE")
    df["Security Code"] = df["Security Code"].astype(str).str.strip()
    df = df[["Security Code", "Industry New Name", "Igroup Name",
             "ISubgroup Name"]].drop_duplicates("Security Code")
    return df.set_index("Security Code")[["Industry New Name", "Igroup Name", "ISubgroup Name"]].to_dict("index")


# ══════════════════════════════════════════════════════════════════════════
# SCREENER
# ══════════════════════════════════════════════════════════════════════════

def screener_login():
    session = requests.Session()
    session.headers.update({
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
        "Referer": BASE_URL,
    })
    r = session.get(f"{BASE_URL}/login/", timeout=30)
    csrf = re.search(r'name="csrfmiddlewaretoken"\s+value="([^"]+)"', r.text)
    if not csrf:
        raise RuntimeError("Could not get CSRF token from Screener login page")
    session.post(f"{BASE_URL}/login/", data={
        "csrfmiddlewaretoken": csrf.group(1),
        "username": SCREENER_EMAIL,
        "password": SCREENER_PASSWORD,
    }, headers={"Referer": f"{BASE_URL}/login/"}, timeout=30)
    check = session.get(f"{BASE_URL}/", timeout=20)
    if SCREENER_EMAIL.split("@")[0].upper() not in check.text.upper() and "ASUTOSH" not in check.text.upper():
        print("  WARNING: Screener login may have failed")
    else:
        print("  Screener login successful")
    return session


def get_ticker_from_bse(session, bse_code):
    try:
        r = session.get(
            f"{BASE_URL}/api/company/search/?q={bse_code}",
            headers={"X-Requested-With": "XMLHttpRequest"},
            timeout=10
        )
        results = r.json()
        if results:
            url = results[0].get("url", "")
            m = re.search(r'/company/([^/]+)/', url)
            if m:
                return m.group(1)
    except Exception:
        pass
    return None


def get_mcap_from_page(session, ticker):
    try:
        r = session.get(f"{BASE_URL}/company/{ticker}/", timeout=15)
        m = re.search(
            r'Market Cap.*?<span class="number">([\d,]+(?:\.\d+)?)</span>\s*\n?\s*Cr\.',
            r.text, re.DOTALL
        )
        if m:
            return float(m.group(1).replace(",", ""))
        m2 = re.search(r'Mkt Cap:\s*([\d,]+(?:\.\d+)?)\s*Crore', r.text)
        if m2:
            return float(m2.group(1).replace(",", ""))
    except Exception:
        pass
    return None


def fetch_market_caps(bse_codes):
    print("  Logging into Screener...")
    try:
        session = screener_login()
    except Exception as e:
        print(f"  Screener login failed: {e}")
        return {}

    codes = list(dict.fromkeys(str(c) for c in bse_codes if pd.notna(c)))
    result = {}
    total = len(codes)

    for i, code in enumerate(codes, 1):
        ticker = get_ticker_from_bse(session, code)
        if ticker:
            mcap = get_mcap_from_page(session, ticker)
            if mcap is not None:
                result[code] = mcap
        if i % 20 == 0 or i == total:
            print(f"  Progress: {i}/{total} | Found: {len(result)}")
        time.sleep(0.3)

    return result


# ══════════════════════════════════════════════════════════════════════════
# EXCEL
# ══════════════════════════════════════════════════════════════════════════

def _hdr(ws, row, ncols, bg=NAVY):
    fill = PatternFill("solid", fgColor=bg)
    font = Font(bold=True, color=WHITE, name="Arial", size=10)
    al = Alignment(horizontal="center", vertical="center", wrap_text=True)
    bd = Side(style="thin", color="FFFFFF")
    bdr = Border(left=bd, right=bd, top=bd, bottom=bd)
    for c in range(1, ncols+1):
        cell = ws.cell(row=row, column=c)
        cell.fill, cell.font, cell.alignment, cell.border = fill, font, al, bdr


def _row(ws, row, ncols, even):
    fill = PatternFill("solid", fgColor=LGRAY if even else WHITE)
    font = Font(name="Arial", size=9)
    al = Alignment(vertical="center")
    bd = Side(style="thin", color=MGRAY)
    bdr = Border(left=bd, right=bd, top=bd, bottom=bd)
    for c in range(1, ncols+1):
        cell = ws.cell(row=row, column=c)
        cell.fill, cell.font, cell.alignment, cell.border = fill, font, al, bdr


def write_sheet(wb, name, df, cols, widths):
    ws = wb.create_sheet(name)
    last = get_column_letter(len(cols))
    ws.merge_cells(f"A1:{last}1")
    tc = ws["A1"]
    tc.value = name
    tc.font = Font(bold=True, color=WHITE, name="Arial", size=12)
    tc.fill = PatternFill("solid", fgColor=BLUE)
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 25
    for ci, col in enumerate(cols, 1):
        ws.cell(row=2, column=ci, value=col)
    _hdr(ws, 2, len(cols))
    ws.row_dimensions[2].height = 28
    for ri, (_, rd) in enumerate(df.iterrows()):
        dr = 3 + ri
        for ci, col in enumerate(cols, 1):
            val = rd.get(col, "")
            if isinstance(val, pd.Timestamp):
                val = val.strftime("%d %b %Y")
            elif not isinstance(val, str) and pd.isna(val):
                val = ""
            ws.cell(row=dr, column=ci, value=val)
        _row(ws, dr, len(cols), even=(ri % 2 == 0))
    ws.freeze_panes = "A3"
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.auto_filter.ref = f"A2:{last}{2+len(df)}"


# ══════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════

def main():
    if not os.path.exists(CSV_IN):
        sys.exit(f"ERROR: {CSV_IN} not found.")

    # Load industry classification
    print("Loading industry classification from 6.0 file...")
    industry_map = load_industry_map()
    print(f"  Loaded {len(industry_map)} BSE codes from Industry - BSE sheet")

    # Load CSV
    df = pd.read_csv(CSV_IN)
    df.columns = [c.strip() for c in df.columns]
    df = df.rename(columns={"Company name": "Company Name"})
    df = df[["Security Code", "Company Name", "Industry", "Purpose",
             "Meeting Date", "Announcement Date"]].dropna(subset=["Security Code"])
    df["Security Code"] = df["Security Code"].astype(str).str.strip()
    df["Meeting Date"] = pd.to_datetime(
        df["Meeting Date"], format="%d %b %Y", errors="coerce")
    df["Announcement Date"] = pd.to_datetime(
        df["Announcement Date"], dayfirst=True, errors="coerce")
    df = df.sort_values("Meeting Date")
    print(f"Loaded {len(df)} board meetings")

    # Attach new industry classification to master
    df["Industry New Name"] = df["Security Code"].map(
        lambda c: industry_map.get(c, {}).get("Industry New Name", ""))
    df["Igroup Name"] = df["Security Code"].map(
        lambda c: industry_map.get(c, {}).get("Igroup Name", ""))
    df["ISubgroup Name"] = df["Security Code"].map(
        lambda c: industry_map.get(c, {}).get("ISubgroup Name", ""))

    # Filter results
    mask = df["Purpose"].apply(lambda p: any(
        kw in str(p).lower() for kw in RESULTS_KEYWORDS))
    df_res = df[mask].copy()
    print(f"Results announcements: {len(df_res)}")

    # Market caps
    print("Fetching market caps from Screener...")
    caps = fetch_market_caps(df_res["Security Code"].tolist())
    df_res["Market Cap (Cr)"] = pd.to_numeric(
        df_res["Security Code"].map(caps), errors="coerce")
    filled = df_res["Market Cap (Cr)"].notna().sum()
    print(f"Market caps fetched: {filled}/{len(df_res)}")

    # Apply 1000 Cr MCap floor
    df_res_filtered = df_res[df_res["Market Cap (Cr)"] >= MCAP_FLOOR].copy()
    print(
        f"After >= {MCAP_FLOOR} Cr MCap filter: {len(df_res_filtered)} companies")

    # Build Excel
    wb = Workbook()
    del wb["Sheet"]

    # Sheet 1: Master — all meetings (no mcap filter)
    write_sheet(wb, "Master - All Meetings", df,
                ["Security Code", "Company Name", "Industry New Name", "Igroup Name", "ISubgroup Name",
                 "Purpose", "Meeting Date", "Announcement Date"],
                [14, 32, 32, 22, 22, 50, 16, 18])

    # Sheet 2: Results by Date & MCap (>=1000 Cr)
    df_r1 = df_res_filtered.sort_values(
        ["Meeting Date", "Market Cap (Cr)"], ascending=[True, False])
    write_sheet(wb, "Results - By Date & MCap", df_r1,
                ["Meeting Date", "Market Cap (Cr)", "Security Code", "Company Name",
                 "Industry New Name", "Igroup Name", "ISubgroup Name", "Purpose"],
                [16, 16, 14, 32, 32, 22, 22, 45])

    # Sheet 3: Results by Sector & Date (>=1000 Cr)
    df_r2 = df_res_filtered.sort_values(
        ["Industry New Name", "Igroup Name", "Meeting Date"])
    write_sheet(wb, "Results - By Sector & Date", df_r2,
                ["Industry New Name", "Igroup Name", "ISubgroup Name",
                 "Meeting Date", "Market Cap (Cr)", "Security Code", "Company Name", "Purpose"],
                [32, 22, 22, 16, 16, 14, 32, 45])

    wb.save(EXCEL_OUT)
    print(f"\n✓ Saved: {EXCEL_OUT}")
    print(f"  Market caps filled: {filled}/{len(df_res)}")
    print(
        f"  Companies after >= {MCAP_FLOOR} Cr filter: {len(df_res_filtered)}")


if __name__ == "__main__":
    main()
