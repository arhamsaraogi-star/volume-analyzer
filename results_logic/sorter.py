"""
Results Tracker - runs daily at 9 AM
--------------------------------------
For each company that published results (T-1):
  1. Scrapes /results/latest/ for the list + basic data
  2. Visits each company's individual page for full quarterly & annual data
  3. Computes QoQ%, YoY Q%, and (if Q4/March) YoY Annual%
  4. Computes EBITDA margin, PAT margin + QoQ/YoY/Annual changes
  5. Enriches with Sector from cache -> BM report -> company page
  6. Sector sheets accumulate ALL companies (no filter)
  7. Sector Summary & Dashboard use MCap >= 1000 Cr companies only

Output workbooks (never lose old data):
  Results_By_Date.xlsx    - one sheet per date, 4 stacked sorted tables
                            (MCap >= 1000 Cr filtered)
  Results_By_Sector.xlsx  - one sheet per sector, all-time accumulation
                            of ALL companies
                          + "Sector Summary" sheet: averages computed
                            from MCap >= 1000 Cr companies only
  Sector_Dashboard.html   - standalone interactive HTML dashboard
                            (MCap >= 1000 Cr averages only)

Usage:
  py sorter.py                # yesterday
  py sorter.py 2026-04-14     # specific date
"""

import requests
import re
import time
import os
import sys
import json
from datetime import datetime, timedelta
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DIR = os.path.dirname(os.path.abspath(__file__))
BM_REPORT = os.path.join(DIR, "Board_Meetings_Report.xlsx")
CACHE_FILE = os.path.join(DIR, "sectors_cache.json")
DATE_WB_PATH = os.path.join(DIR, "Results_By_Date.xlsx")
SECTOR_WB_PATH = os.path.join(DIR, "Results_By_Sector.xlsx")
DASHBOARD_HTML = os.path.join(DIR, "Sector_Dashboard.html")
DAILY_DASHBOARD_HTML = os.path.join(DIR, "Daily_Dashboard.html")

SCREENER_EMAIL = "asutosh@ashikagroup.com"
SCREENER_PASSWORD = "Dilipsir@1234"
BASE_URL = "https://www.screener.in"

MCAP_FLOOR = 0  # Cr - lowered to 0 to rely completely on UI Range filter

# -- Subsector lookup (from file 6.0) ----------------------------------------
_SUBSECTOR_LOOKUP_PATH = os.path.join(DIR, "subsector_lookup.json")
try:
    with open(_SUBSECTOR_LOOKUP_PATH, encoding="utf-8") as _f:
        SUBSECTOR_LOOKUP = json.load(_f)
except Exception:
    SUBSECTOR_LOOKUP = {}

# -- Indices lookup (from screener dynamic scrape) ---------------------------
_INDICES_LOOKUP_PATH = os.path.join(DIR, "indices_lookup.json")
try:
    with open(_INDICES_LOOKUP_PATH, encoding="utf-8") as _f:
        INDICES_LOOKUP = json.load(_f)
except Exception:
    INDICES_LOOKUP = {}

# -- Styling ----------------------------------------------------------------
NAVY = "0E2841"
BLUE = "243F8D"
WHITE = "FFFFFF"
LGRAY = "F2F2F2"
MGRAY = "D9D9D9"
GREEN = "00B050"
RED = "C00000"
AMBER = "FFC000"
SECTION_COLORS = ["1F4E79", "243F8D", "2E75B6", "2F5496"]

# -- Column definition ------------------------------------------------------
Q_COLS = [
    "Result Date", "Company Name", "Sector", "Industry Group", "Subsector", "Indices", "Market Cap (Cr)", "Price", "PE", "Screener Ticker",
    "Quarter",
    "Sales (Cr)", "EBITDA (Cr)", "Net Profit (Cr)", "EPS (Rs)",
    "EBITDA Margin%", "PAT Margin%",
    "Sales QoQ%", "EBITDA QoQ%", "NP QoQ%", "EPS QoQ%",
    "EBITDA Margin QoQ pp", "PAT Margin QoQ pp",
    "Sales YoY Q%", "EBITDA YoY Q%", "NP YoY Q%", "EPS YoY Q%",
    "EBITDA Margin YoY pp", "PAT Margin YoY pp",
    "Prev Qtr Sales", "Prev Qtr EBITDA", "Prev Qtr NP", "Prev Qtr EPS",
    "LY Qtr Sales", "LY Qtr EBITDA", "LY Qtr NP", "LY Qtr EPS",
]

A_COLS = [
    "FY Sales (Cr)", "FY EBITDA (Cr)", "FY NP (Cr)", "FY EPS (Rs)",
    "FY EBITDA Margin%", "FY PAT Margin%",
    "FY Sales YoY%", "FY EBITDA YoY%", "FY NP YoY%", "FY EPS YoY%",
    "FY EBITDA Margin YoY pp", "FY PAT Margin YoY pp",
    "Prev FY Sales", "Prev FY EBITDA", "Prev FY NP", "Prev FY EPS",
]

ALL_COLS = Q_COLS + A_COLS

COL_WIDTHS = [
    14, 28, 28, 22, 28, 16, 10, 8, 16, 12,
    13, 13, 14, 10, 14, 12,
    12, 12, 10, 10, 16, 14,
    13, 13, 11, 11, 16, 14,
    16, 16, 14, 14,
    14, 14, 12, 12,
    14, 14, 12, 10, 14, 12,
    14, 14, 12, 10, 16, 14,
    14, 14, 12, 12,
]

YOY_COLS = {
    "Sales YoY Q%", "EBITDA YoY Q%", "NP YoY Q%", "EPS YoY Q%",
    "FY Sales YoY%", "FY EBITDA YoY%", "FY NP YoY%", "FY EPS YoY%",
}
QOQ_COLS = {"Sales QoQ%", "EBITDA QoQ%", "NP QoQ%", "EPS QoQ%"}
MARGIN_DELTA_COLS = {
    "EBITDA Margin QoQ pp", "PAT Margin QoQ pp",
    "EBITDA Margin YoY pp", "PAT Margin YoY pp",
    "FY EBITDA Margin YoY pp", "FY PAT Margin YoY pp",
}
PCT_COLS = YOY_COLS | QOQ_COLS | MARGIN_DELTA_COLS

SORT_COLS = ["Sales YoY Q%", "EBITDA YoY Q%", "NP YoY Q%", "EPS YoY Q%"]

# Sector summary metrics
SUMMARY_METRICS = [
    ("Avg Sales YoY Q%",          "Sales YoY Q%"),
    ("Avg EBITDA YoY Q%",         "EBITDA YoY Q%"),
    ("Avg NP YoY Q%",             "NP YoY Q%"),
    ("Avg EPS YoY Q%",            "EPS YoY Q%"),
    ("Avg EBITDA Margin%",        "EBITDA Margin%"),
    ("Avg PAT Margin%",           "PAT Margin%"),
    ("Avg EBITDA Margin YoY pp",  "EBITDA Margin YoY pp"),
    ("Avg PAT Margin YoY pp",     "PAT Margin YoY pp"),
    ("Avg EBITDA Margin QoQ pp",  "EBITDA Margin QoQ pp"),
    ("Avg PAT Margin QoQ pp",     "PAT Margin QoQ pp"),
    ("Avg FY Sales YoY%",         "FY Sales YoY%"),
    ("Avg FY EBITDA YoY%",        "FY EBITDA YoY%"),
    ("Avg FY NP YoY%",            "FY NP YoY%"),
    ("Avg FY EPS YoY%",           "FY EPS YoY%"),
    ("Avg FY EBITDA Margin%",     "FY EBITDA Margin%"),
    ("Avg FY PAT Margin%",        "FY PAT Margin%"),
    ("Avg FY EBITDA Margin YoY pp", "FY EBITDA Margin YoY pp"),
    ("Avg FY PAT Margin YoY pp",  "FY PAT Margin YoY pp"),
]


# ══════════════════════════════════════════════════════════════════════════
# MCAP FILTER (replaces Nifty 500)
# ══════════════════════════════════════════════════════════════════════════

def filter_by_mcap(df, floor=MCAP_FLOOR):
    if "Market Cap (Cr)" not in df.columns:
        print(f"  WARNING: Market Cap column missing - no MCap filter applied")
        return df
    mcap_col = pd.to_numeric(df["Market Cap (Cr)"], errors="coerce")
    result = df[mcap_col >= floor].copy()
    print(
        f"  MCap >= {floor} Cr filter: kept {len(result)}/{len(df)} (removed {len(df)-len(result)})")
    return result


# ══════════════════════════════════════════════════════════════════════════
# HELPERS
# ══════════════════════════════════════════════════════════════════════════

def pn(txt):
    if txt is None:
        return None
    txt = re.sub(r'[⇡⇣↑↓,₹%\s]', '', str(txt)).strip()
    try:
        return float(txt)
    except:
        return None


def pct(new, old):
    """
    Percentage growth from old to new.
    Returns None (shown as dash) if:
      - either value is missing
      - old is zero (division by zero)
      - old is negative: negative-to-positive turnaround (e.g. -4 -> +4)
      - new is negative: positive-to-negative deterioration (e.g. +122 -> -118)
    Both sign-crossing directions are incomparable as percentages.
    """
    if new is None or old is None or old == 0:
        return None
    if old < 0 or new < 0:
        # Either direction of sign-crossing makes % growth meaningless
        return None
    return round((new - old) / abs(old) * 100, 1)



def margin(ebitda, sales):
    if ebitda is None or sales is None or sales == 0:
        return None
    return round(ebitda / sales * 100, 2)


def pp_delta(m_new, m_old):
    """Percentage point delta between two margins."""
    if m_new is None or m_old is None:
        return None
    return round(m_new - m_old, 2)


def safe_name(s):
    return re.sub(r'[\\/*?:\[\]]', '', str(s))[:31]


def load_or_new(path):
    if os.path.exists(path):
        return load_workbook(path)
    wb = Workbook()
    del wb["Sheet"]
    return wb


# ══════════════════════════════════════════════════════════════════════════
# SECTOR CACHE
# ══════════════════════════════════════════════════════════════════════════

def load_cache():
    if os.path.exists(CACHE_FILE):
        with open(CACHE_FILE) as f:
            return json.load(f)
    return {}


def save_cache(cache):
    with open(CACHE_FILE, "w") as f:
        json.dump(cache, f, indent=2)


def build_bm_lookup(cache):
    """Read Board_Meetings_Report.xlsx and pull Industry New Name as sector."""
    if not os.path.exists(BM_REPORT):
        return cache
    try:
        df = pd.ExcelFile(BM_REPORT).parse(
            "Results - By Date & MCap", header=1)
        df.columns = [str(c).strip() for c in df.columns]
        # Prefer Industry New Name if available, fall back to Industry
        sector_col = "Industry New Name" if "Industry New Name" in df.columns else "Industry"
        for _, row in df.iterrows():
            t = str(row.get("Company Name", "")).strip().upper()
            s = str(row.get(sector_col, "")).strip()
            m = row.get("Market Cap (Cr)", None)
            if t and s and s != "nan" and t not in cache:
                cache[t] = {"sector": s, "mcap": m}
        print(f"  BM lookup done -> cache {len(cache)} tickers")
    except Exception as e:
        print(f"  WARNING BM lookup: {e}")
    return cache


# ══════════════════════════════════════════════════════════════════════════
# SCREENER LOGIN
# ══════════════════════════════════════════════════════════════════════════

def screener_login():
    s = requests.Session()
    s.headers.update(
        {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"})
    r = s.get(f"{BASE_URL}/login/", timeout=30)
    csrf = re.search(r'name="csrfmiddlewaretoken"\s+value="([^"]+)"', r.text)
    if not csrf:
        raise RuntimeError("Cannot get CSRF")
    s.post(f"{BASE_URL}/login/",
           headers={"Referer": f"{BASE_URL}/login/"},
           data={"csrfmiddlewaretoken": csrf.group(1),
                 "username": SCREENER_EMAIL, "password": SCREENER_PASSWORD},
           timeout=30)
    print("  Screener login OK")
    return s


# ══════════════════════════════════════════════════════════════════════════
# SCRAPE /results/latest/
# ══════════════════════════════════════════════════════════════════════════

def parse_yoy_span(td):
    if not td:
        return None
    span = td.find("span", class_="change")
    if not span:
        return None
    txt = re.sub(r'[⇡⇣↑↓%]', '', span.get_text(strip=True)).strip()
    try:
        val = float(txt)
        if "down" in span.get("class", []):
            val = -abs(val)
        return val
    except:
        return None


def scrape_results_list(session, dt):
    d, mo, y = dt.day, dt.month, dt.year
    base = (f"{BASE_URL}/results/latest/"
            f"?result_update_date__day={d}&result_update_date__month={mo}"
            f"&result_update_date__year={y}")
    companies = []
    page = 1
    while True:
        url = base + (f"&p={page}" if page > 1 else "")
        r = session.get(url, timeout=30)
        soup = BeautifulSoup(r.text, "html.parser")
        hdrs = soup.find_all("div", class_=lambda c: c and
                             "flex-row" in c and "margin-top-32" in c)
        if not hdrs:
            break
        for hdr in hdrs:
            rec = {}
            a = hdr.find("a", class_="font-weight-500")
            if not a:
                continue
            rec["Company Name"] = a.get_text(" ", strip=True)
            m = re.search(r'/company/([^/]+)/', a.get("href", ""))
            rec["_ticker"] = m.group(1).upper() if m else ""
            rec["_url"] = a.get("href", "")

            meta = hdr.find("div", class_="font-size-14")
            if meta:
                for sp in meta.find_all("span", class_="sub"):
                    txt = sp.get_text(" ", strip=True)
                    strong = sp.find("span", class_="strong")
                    val = pn(strong.get_text() if strong else "")
                    if "Price" in txt:
                        rec["Price"] = val
                    elif sp.has_attr("data-mcap") or "M.Cap" in txt:
                        rec["Market Cap (Cr)"] = val
                    elif "PE" in txt and "M.Cap" not in txt and "Price" not in txt:
                        rec["PE"] = val

            tbl = hdr.find_next_sibling("div")
            if tbl:
                table = tbl.find("table")
                if table:
                    ths = [th.get_text(strip=True)
                           for th in table.find_all("th")]
                    rec["Quarter"] = ths[2] if len(ths) > 2 else ""

                    def gr(attr):
                        row = table.find("tr", attrs={attr: True})
                        if not row:
                            return None
                        tds = row.find_all("td")
                        return parse_yoy_span(tds[1]) if len(tds) > 1 else None

                    rec["_screener_sales_yoy"] = gr("data-sales")
                    rec["_screener_np_yoy"] = gr("data-net-profit")
                    rec["_screener_eps_yoy"] = gr("data-eps")
            companies.append(rec)

        pager = soup.find("p", class_="paginator")
        has_next = pager and any(
            a.get_text(strip=True).lower() in ("next", "›", "»")
            for a in pager.find_all("a"))
        if not has_next:
            break
        page += 1
        time.sleep(0.5)

    print(f"  Found {len(companies)} companies on results page")
    return companies


# ══════════════════════════════════════════════════════════════════════════
# SCRAPE INDIVIDUAL COMPANY PAGE
# ══════════════════════════════════════════════════════════════════════════

def get_table_rows(soup, section_id):
    section = soup.find("section", id=section_id)
    if not section:
        return {}, []
    holder = section.find("div", attrs={"data-result-table": True})
    if not holder:
        return {}, []
    table = holder.find("table")
    if not table:
        return {}, []

    ths = table.find("thead").find_all("th")
    headers = [th.get_text(strip=True) for th in ths[1:]]

    # -- Drop the TTM column if present ----------------------------------
    # Screener appends a "TTM" column at the end of quarterly tables.
    # This column must be excluded so that index -1 = latest actual quarter,
    # index -2 = previous quarter, and index -5 = same quarter last year.
    ttm_indices = set()
    for i, h in enumerate(headers):
        if h.strip().upper() == "TTM":
            ttm_indices.add(i)
    if ttm_indices:
        headers = [h for i, h in enumerate(headers) if i not in ttm_indices]

    rows = {}
    for tr in table.find("tbody").find_all("tr"):
        label_td = tr.find("td", class_="text")
        if not label_td:
            continue
        label = label_td.get_text(strip=True).lower()
        label = re.sub(r'\+\s*$', '', label).strip()
        all_tds = tr.find_all("td")[1:]
        vals = []
        for i, td in enumerate(all_tds):
            if i in ttm_indices:
                continue  # skip the TTM column values
            txt = td.get_text(strip=True).replace(",", "").replace("%", "")
            try:
                vals.append(float(txt))
            except:
                vals.append(None)
        rows[label] = vals
    return rows, headers


def scrape_company_page(session, ticker, url, cache):
    rec = {}
    base_path = re.sub(r'#.*$', '', url.strip())
    if not base_path.endswith('/'):
        base_path += '/'
    full_url = BASE_URL + base_path

    try:
        r = session.get(full_url, timeout=20)
        soup = BeautifulSoup(r.text, "html.parser")
    except Exception as e:
        print(f"    WARNING: Cannot fetch {full_url}: {e}")
        return rec

    sector = None
    for a in soup.find_all("a"):
        if a.get("title") == "Industry":
            sector = a.get_text(strip=True)
            break
    if not sector:
        for a in soup.find_all("a"):
            if a.get("title") in ("Sector", "Broad Industry"):
                sector = a.get_text(strip=True)
                break
    if sector:
        rec["Sector"] = sector
        if ticker and ticker not in cache:
            cache[ticker] = {"sector": sector, "mcap": None}
        elif ticker and not cache.get(ticker, {}).get("sector"):
            cache.setdefault(ticker, {})["sector"] = sector

    for li in soup.find_all("li", attrs={"data-source": True}):
        ns = li.find("span", class_="name")
        if ns and "Market Cap" in ns.get_text():
            num = li.find("span", class_="number")
            if num:
                try:
                    rec["_page_mcap"] = float(
                        num.get_text(strip=True).replace(",", ""))
                except:
                    pass
            break

    qrows, qhdrs = get_table_rows(soup, "quarters")
    if not qhdrs:
        return rec

    n = len(qhdrs)
    if n < 1:
        return rec

    latest_qtr = qhdrs[-1]
    rec["Quarter"] = latest_qtr

    def find_row(rows, *keys):
        for k in keys:
            for label in rows:
                if k.lower() in label.lower():
                    return rows[label]
        return None

    def v(row, idx):
        # Safe index access - returns None if row missing or index out of range
        if row is None:
            return None
        try:
            return row[idx]
        except IndexError:
            return None

    def find_yoy_idx(hdrs):
        """
        Return the list-index (negative) of the quarter that is exactly
        1 year before hdrs[-1], matched by name (e.g. 'Mar 2025' for 'Mar 2026').
        Falls back to -5 only if the matching header is not found.
        This fixes the bug where missing quarters (e.g. Jun 2025 not filed)
        caused the code to pick the wrong base (Dec 2024 instead of Mar 2025).
        """
        if not hdrs:
            return -5
        latest = hdrs[-1].strip()
        m = re.match(r'(\w{3})\s+(\d{4})', latest)
        if not m:
            return -5
        month_str, year = m.group(1), int(m.group(2))
        target = f"{month_str} {year - 1}"   # e.g. 'Mar 2025'
        for i, h in enumerate(hdrs):
            if h.strip().startswith(target):
                idx = i - len(hdrs)  # convert to negative index
                print(f"    YoY base: '{h.strip()}' (index {idx}) for latest '{latest}'")
                return idx
        # Not found - log warning and fall back
        print(f"    WARNING: YoY base quarter '{target}' not in headers {hdrs} - using None")
        return None  # will cause v() to return None safely

    # Sales: financial/insurance companies use different labels
    sales_r = find_row(qrows,
                       "sales", "revenue from operations", "net premium earned",
                       "total income", "net interest income", "revenue")
    # EBITDA: some companies omit this row entirely
    ebitda_r = find_row(qrows,
                        "operating profit", "ebitda", "profit before interest and tax",
                        "profit before exceptional")
    np_r = find_row(qrows, "net profit", "profit after tax", "pat")
    eps_r = find_row(qrows, "eps in rs", "eps")

    # Current quarter values
    s0 = v(sales_r,  -1)
    e0 = v(ebitda_r, -1)
    n0 = v(np_r,     -1)
    ep0 = v(eps_r,    -1)

    # Previous quarter (QoQ)
    s1 = v(sales_r,  -2)
    e1 = v(ebitda_r, -2)
    n1 = v(np_r,     -2)
    ep1 = v(eps_r,    -2)

    # Last year same quarter (YoY) -- matched by quarter name, not fixed index.
    # Using a fixed index like -5 breaks when quarters are missing (e.g. a
    # company that didn't file Jun 2025 results causes the 5-column table to
    # contain Dec24/Mar25/Sep25/Dec25/Mar26 -- index -5 = Dec 2024 (WRONG)
    # instead of Mar 2025 (CORRECT), inflating YoY from ~100% to ~400%.
    yoy_idx = find_yoy_idx(qhdrs)
    if yoy_idx is None:
        s4 = e4 = n4 = ep4 = None
    else:
        s4  = v(sales_r,  yoy_idx)
        e4  = v(ebitda_r, yoy_idx)
        n4  = v(np_r,     yoy_idx)
        ep4 = v(eps_r,    yoy_idx)

    rec["Sales (Cr)"] = s0
    rec["EBITDA (Cr)"] = e0
    rec["Net Profit (Cr)"] = n0
    rec["EPS (Rs)"] = ep0

    rec["Prev Qtr Sales"] = s1
    rec["Prev Qtr EBITDA"] = e1
    rec["Prev Qtr NP"] = n1
    rec["Prev Qtr EPS"] = ep1

    rec["LY Qtr Sales"] = s4
    rec["LY Qtr EBITDA"] = e4
    rec["LY Qtr NP"] = n4
    rec["LY Qtr EPS"] = ep4

    # Margins current quarter
    m0_ebitda = margin(e0, s0)
    m0_pat = margin(n0, s0)
    m1_ebitda = margin(e1, s1)
    m1_pat = margin(n1, s1)
    m4_ebitda = margin(e4, s4)
    m4_pat = margin(n4, s4)

    rec["EBITDA Margin%"] = m0_ebitda
    rec["PAT Margin%"] = m0_pat
    rec["EBITDA Margin QoQ pp"] = pp_delta(m0_ebitda, m1_ebitda)
    rec["PAT Margin QoQ pp"] = pp_delta(m0_pat,    m1_pat)
    rec["EBITDA Margin YoY pp"] = pp_delta(m0_ebitda, m4_ebitda)
    rec["PAT Margin YoY pp"] = pp_delta(m0_pat,    m4_pat)

    # Growth rates
    rec["Sales QoQ%"] = pct(s0, s1)
    rec["EBITDA QoQ%"] = pct(e0, e1)
    rec["NP QoQ%"] = pct(n0, n1)
    rec["EPS QoQ%"] = pct(ep0, ep1)

    rec["Sales YoY Q%"] = pct(s0, s4)
    rec["EBITDA YoY Q%"] = pct(e0, e4)
    rec["NP YoY Q%"] = pct(n0, n4)
    rec["EPS YoY Q%"] = pct(ep0, ep4)

    is_q4 = latest_qtr.startswith("Mar")
    if is_q4:
        arows, ahdrs = get_table_rows(soup, "profit-loss")
        if ahdrs and len(ahdrs) >= 2:
            as_r = find_row(arows,
                            "sales", "revenue from operations", "net premium earned",
                            "total income", "net interest income", "revenue")
            ae_r = find_row(arows,
                            "operating profit", "ebitda", "profit before interest and tax",
                            "profit before exceptional")
            anp_r = find_row(arows, "net profit", "profit after tax", "pat")
            aep_r = find_row(arows, "eps in rs", "eps")

            fy_s = v(as_r,  -1)
            fy_e = v(ae_r,  -1)
            fy_n = v(anp_r, -1)
            fy_ep = v(aep_r, -1)
            pfy_s = v(as_r,  -2)
            pfy_e = v(ae_r,  -2)
            pfy_n = v(anp_r, -2)
            pfy_ep = v(aep_r, -2)

            rec["FY Sales (Cr)"] = fy_s
            rec["FY EBITDA (Cr)"] = fy_e
            rec["FY NP (Cr)"] = fy_n
            rec["FY EPS (Rs)"] = fy_ep

            rec["Prev FY Sales"] = pfy_s
            rec["Prev FY EBITDA"] = pfy_e
            rec["Prev FY NP"] = pfy_n
            rec["Prev FY EPS"] = pfy_ep

            fy_em = margin(fy_e,  fy_s)
            fy_pm = margin(fy_n,  fy_s)
            pfy_em = margin(pfy_e, pfy_s)
            pfy_pm = margin(pfy_n, pfy_s)

            rec["FY EBITDA Margin%"] = fy_em
            rec["FY PAT Margin%"] = fy_pm
            rec["FY EBITDA Margin YoY pp"] = pp_delta(fy_em, pfy_em)
            rec["FY PAT Margin YoY pp"] = pp_delta(fy_pm, pfy_pm)

            rec["FY Sales YoY%"] = pct(fy_s,  pfy_s)
            rec["FY EBITDA YoY%"] = pct(fy_e,  pfy_e)
            rec["FY NP YoY%"] = pct(fy_n,  pfy_n)
            rec["FY EPS YoY%"] = pct(fy_ep, pfy_ep)

    return rec


# ══════════════════════════════════════════════════════════════════════════
# ENRICH ALL COMPANIES
# ══════════════════════════════════════════════════════════════════════════

def enrich_all(session, companies, cache, date_str):
    total = len(companies)
    records = []

    for i, base in enumerate(companies, 1):
        ticker = base.get("_ticker", "")
        url = base.get("_url", "")
        print(f"  [{i}/{total}] {base['Company Name']} ({ticker})")

        # Retry once on empty result (handles transient timeouts / rate-limits)
        page_data = scrape_company_page(session, ticker, url, cache)
        if not any(k in page_data for k in ("Sales (Cr)", "Net Profit (Cr)", "Quarter")):
            print(f"    Retrying {base['Company Name']}...")
            time.sleep(3)
            page_data = scrape_company_page(session, ticker, url, cache)
        rec = {**base, **page_data}

        if not rec.get("Sector"):
            rec["Sector"] = cache.get(ticker, {}).get(
                "sector", "Other") or "Other"

        if not rec.get("Market Cap (Cr)") and rec.get("_page_mcap"):
            rec["Market Cap (Cr)"] = rec["_page_mcap"]
        if not rec.get("Market Cap (Cr)"):
            rec["Market Cap (Cr)"] = cache.get(ticker, {}).get("mcap")

        rec["Result Date"] = date_str

        ticker_key = rec.get("_ticker", "") or ""
        rec["Screener Ticker"] = ticker_key
        # Subsector enrichment from file 6.0 lookup
        sub_info = SUBSECTOR_LOOKUP.get(ticker_key.upper(), {})
        
        # Override Sector using "Industry New Name" (if available) from the BSE lookup
        new_sector = sub_info.get("industry", "")
        if new_sector:
            rec["Sector"] = new_sector
            
        rec["Industry Group"] = new_sector
        rec["Subsector"] = sub_info.get("sub", "")
        
        # Map dynamic indices
        idx_list = INDICES_LOOKUP.get(ticker_key.upper(), [])
        idx_str = ", ".join(idx_list) if idx_list else ""
        rec["Indices"] = idx_str
        # print(f"      Mapped indices for {ticker_key}: {idx_str}")
        for k in ["_ticker", "_url", "_page_mcap",
                  "_screener_sales_yoy", "_screener_np_yoy", "_screener_eps_yoy"]:
            rec.pop(k, None)

        records.append(rec)
        time.sleep(0.4)

    return records


# ══════════════════════════════════════════════════════════════════════════
# EXCEL STYLING
# ══════════════════════════════════════════════════════════════════════════

def _border(color=MGRAY):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def style_section_hdr(ws, row, ncols, label, bg):
    last = get_column_letter(ncols)
    ws.merge_cells(f"A{row}:{last}{row}")
    c = ws[f"A{row}"]
    c.value = label
    c.font = Font(bold=True, color=WHITE, name="Arial", size=11)
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 22


def style_col_hdr(ws, row, ncols, bg):
    fill = PatternFill("solid", fgColor=bg)
    font = Font(bold=True, color=WHITE, name="Arial", size=9)
    al = Alignment(horizontal="center", vertical="center", wrap_text=True)
    bdr = _border("FFFFFF")
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill, cell.font, cell.alignment, cell.border = fill, font, al, bdr
    ws.row_dimensions[row].height = 32


def style_data_row(ws, row, ncols, even, col_map):
    bg = LGRAY if even else WHITE
    fill = PatternFill("solid", fgColor=bg)
    bdr = _border()
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        col_name = col_map.get(c, "")
        val = cell.value
        if col_name in PCT_COLS and isinstance(val, (int, float)):
            color = GREEN if val >= 0 else RED
            cell.font = Font(name="Arial", size=9, color=color, bold=True)
        else:
            cell.font = Font(name="Arial", size=9)
        cell.fill = fill
        cell.alignment = Alignment(vertical="center")
        cell.border = bdr


def write_data_rows(ws, df_in, cols, start_row, col_map):
    df_out = df_in.reindex(columns=cols)
    for ri, (_, rd) in enumerate(df_out.iterrows()):
        dr = start_row + ri
        for ci, col in enumerate(cols, 1):
            val = rd.get(col, "")
            if not isinstance(val, str) and pd.isna(val):
                val = ""
            ws.cell(row=dr, column=ci, value=val)
        style_data_row(ws, dr, len(cols), even=(ri % 2 == 0), col_map=col_map)
    return start_row + len(df_out)


# ══════════════════════════════════════════════════════════════════════════
# WRITE DATE SHEET
# ══════════════════════════════════════════════════════════════════════════

def write_date_sheet(ws, date_str, df):
    nc = len(ALL_COLS)
    col_map = {i + 1: c for i, c in enumerate(ALL_COLS)}
    last = get_column_letter(nc)

    ws.merge_cells(f"A1:{last}1")
    tc = ws["A1"]
    tc.value = f"Results  ·  {date_str}"
    tc.font = Font(bold=True, color=WHITE, name="Arial", size=13)
    tc.fill = PatternFill("solid", fgColor=NAVY)
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    current_row = 2
    for idx, sort_col in enumerate(SORT_COLS):
        bg = SECTION_COLORS[idx]
        style_section_hdr(ws, current_row, nc, f"▶  Sorted by {sort_col}", bg)
        current_row += 1
        for ci, col in enumerate(ALL_COLS, 1):
            ws.cell(row=current_row, column=ci, value=col)
        style_col_hdr(ws, current_row, nc, bg)
        current_row += 1
        df_s = df.sort_values(sort_col, ascending=False, na_position="last")
        current_row = write_data_rows(ws, df_s, ALL_COLS, current_row, col_map)
        current_row += 1

    ws.freeze_panes = "A2"
    for ci, w in enumerate(COL_WIDTHS[:nc], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w


# ══════════════════════════════════════════════════════════════════════════
# SECTOR SHEET
# ══════════════════════════════════════════════════════════════════════════

def _write_sector_title(ws, sector, nc):
    last = get_column_letter(nc)
    ws.merge_cells(f"A1:{last}1")
    tc = ws["A1"]
    tc.value = sector
    tc.font = Font(bold=True, color=WHITE, name="Arial", size=13)
    tc.fill = PatternFill("solid", fgColor=NAVY)
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    for ci, col in enumerate(ALL_COLS, 1):
        ws.cell(row=2, column=ci, value=col)
    style_col_hdr(ws, 2, nc, BLUE)

    ws.freeze_panes = "A3"
    for ci, w in enumerate(COL_WIDTHS[:nc], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w


def write_sector_fresh(ws, sector, df):
    nc = len(ALL_COLS)
    col_map = {i + 1: c for i, c in enumerate(ALL_COLS)}
    last = get_column_letter(nc)
    _write_sector_title(ws, sector, nc)
    df_s = df.sort_values("Sales YoY Q%", ascending=False, na_position="last")
    write_data_rows(ws, df_s, ALL_COLS, 3, col_map)
    ws.auto_filter.ref = f"A2:{last}{2 + len(df_s)}"


def read_sheet_df(ws):
    nc = len(ALL_COLS)
    rows = []
    for r in range(3, ws.max_row + 1):
        vals = [ws.cell(row=r, column=c).value for c in range(1, nc + 1)]
        if any(v is not None for v in vals):
            rows.append(dict(zip(ALL_COLS, vals)))
    return pd.DataFrame(rows) if rows else pd.DataFrame(columns=ALL_COLS)


def rewrite_sector_data(ws, df_in):
    nc = len(ALL_COLS)
    col_map = {i + 1: c for i, c in enumerate(ALL_COLS)}
    last = get_column_letter(nc)
    sector = ws["A1"].value or ws.title
    _write_sector_title(ws, str(sector), nc)
    if ws.max_row >= 3:
        ws.delete_rows(3, ws.max_row - 2)
    df_s = df_in.sort_values(
        "Sales YoY Q%", ascending=False, na_position="last")
    write_data_rows(ws, df_s, ALL_COLS, 3, col_map)
    ws.auto_filter.ref = f"A2:{last}{2 + len(df_s)}"


# ══════════════════════════════════════════════════════════════════════════
# SECTOR SUMMARY SHEET
# ══════════════════════════════════════════════════════════════════════════

def _apply_summary_color(ws, row, col, val, is_pp=False):
    if val is None or (isinstance(val, str) and val == ""):
        return
    try:
        v = float(val)
    except:
        return
    if is_pp:
        # Margin pp deltas: +/-1pp thresholds
        if v >= 2:
            color = "00B050"
        elif v >= 0:
            color = "92D050"
        elif v >= -2:
            color = "FFC7CE"
        else:
            color = "C00000"
        font_color = "FFFFFF" if abs(v) >= 2 else "000000"
    else:
        if v >= 15:
            color = "00B050"
        elif v >= 5:
            color = "92D050"
        elif v >= 0:
            color = "FFEB9C"
        elif v >= -10:
            color = "FFC7CE"
        else:
            color = "C00000"
        font_color = "FFFFFF" if v >= 15 or v <= -10 else "000000"
    ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=color)
    ws.cell(row=row, column=col).font = Font(
        name="Arial", size=9, bold=(abs(v) >= 15 if not is_pp else abs(v) >= 2),
        color=font_color)


def build_sector_summary(wb, df_all):
    """
    Build / refresh 'Sector Summary' sheet.
    df_all: all rows from Results_By_Sector.xlsx (MCap >= 1000 Cr filtered).
    """
    SNAME = "Sector Summary"
    if SNAME in wb.sheetnames:
        del wb[SNAME]
    ws = wb.create_sheet(SNAME, 0)

    metric_labels = [m[0] for m in SUMMARY_METRICS]
    metric_cols = [m[1] for m in SUMMARY_METRICS]
    # Track which metrics are pp deltas for coloring
    pp_metrics = {m[0] for m in SUMMARY_METRICS if "pp" in m[0].lower()}

    num_cols = [c for c in ALL_COLS if c not in {
        "Result Date", "Company Name", "Sector", "Quarter",
        "Screener Ticker", "Industry Group", "Subsector"}]

    for c in num_cols:
        if c in df_all.columns:
            df_all[c] = pd.to_numeric(df_all[c], errors="coerce")

    # Benchmark: averages across all MCap >= 1000 Cr companies
    n500_avgs = {}
    for lbl, col in SUMMARY_METRICS:
        if col in df_all.columns:
            n500_avgs[lbl] = round(df_all[col].mean(skipna=True), 1)
        else:
            n500_avgs[lbl] = None

    sectors = sorted(df_all["Sector"].dropna().unique())
    summary_rows = []
    for sec in sectors:
        df_sec = df_all[df_all["Sector"] == sec]
        row = {"Sector": sec, "Companies": len(df_sec)}
        for lbl, col in SUMMARY_METRICS:
            if col in df_sec.columns:
                row[lbl] = round(df_sec[col].mean(skipna=True),
                                 1) if df_sec[col].notna().any() else None
            else:
                row[lbl] = None
        summary_rows.append(row)

    df_sum = pd.DataFrame(summary_rows)

    TOTAL_COLS = 2 + len(metric_labels)
    last_col = get_column_letter(TOTAL_COLS)

    # Row 1: Title
    ws.merge_cells(f"A1:{last_col}1")
    tc = ws["A1"]
    tc.value = f"Sector Summary  ·  MCap >= {MCAP_FLOOR} Cr Universe  ·  Growth & Margins vs Benchmark"
    tc.font = Font(bold=True, color=WHITE, name="Arial", size=14)
    tc.fill = PatternFill("solid", fgColor=NAVY)
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    # Row 2: Column headers
    headers = ["Sector", "Companies"] + metric_labels
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=ci, value=h)
        cell.font = Font(bold=True, color=WHITE, name="Arial", size=9)
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True)
        cell.border = _border("FFFFFF")
    ws.row_dimensions[2].height = 40

    # Row 3: Benchmark row
    brow = ws["A3"]
    brow.value = f"▶  Broad Market Avg (MCap >= {MCAP_FLOOR} Cr)"
    brow.font = Font(bold=True, color=WHITE, name="Arial", size=10)
    brow.fill = PatternFill("solid", fgColor="1F4E79")
    brow.alignment = Alignment(vertical="center")
    brow.border = _border("FFFFFF")

    ws.cell(row=3, column=2, value=len(df_all)).font = Font(
        bold=True, color=WHITE, name="Arial", size=10)
    ws.cell(row=3, column=2).fill = PatternFill("solid", fgColor="1F4E79")
    ws.cell(row=3, column=2).alignment = Alignment(
        horizontal="center", vertical="center")
    ws.cell(row=3, column=2).border = _border("FFFFFF")

    for ci, lbl in enumerate(metric_labels, 3):
        val = n500_avgs.get(lbl)
        cell = ws.cell(row=3, column=ci)
        cell.value = val if val is not None else ""
        cell.font = Font(bold=True, color=WHITE, name="Arial", size=10)
        cell.fill = PatternFill("solid", fgColor="1F4E79")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _border("FFFFFF")
        cell.number_format = '0.0'
    ws.row_dimensions[3].height = 24

    # Data rows
    primary_metric = metric_labels[0]
    if not df_sum.empty and primary_metric in df_sum.columns:
        df_sum_sorted = df_sum.sort_values(
            primary_metric, ascending=False, na_position="last")
    else:
        df_sum_sorted = df_sum

    for ri, (_, rd) in enumerate(df_sum_sorted.iterrows()):
        dr = 4 + ri
        even = (ri % 2 == 0)
        bg = LGRAY if even else WHITE

        sc = ws.cell(row=dr, column=1, value=rd["Sector"])
        sc.font = Font(name="Arial", size=9, bold=True)
        sc.fill = PatternFill("solid", fgColor=bg)
        sc.alignment = Alignment(vertical="center")
        sc.border = _border()

        cc = ws.cell(row=dr, column=2, value=rd["Companies"])
        cc.font = Font(name="Arial", size=9)
        cc.fill = PatternFill("solid", fgColor=bg)
        cc.alignment = Alignment(horizontal="center", vertical="center")
        cc.border = _border()
        cc.number_format = "0"

        for ci, lbl in enumerate(metric_labels, 3):
            val = rd.get(lbl)
            cell = ws.cell(row=dr, column=ci)
            cell.value = val if (val is not None and not (
                isinstance(val, float) and pd.isna(val))) else ""
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = _border()
            cell.number_format = '0.0'
            if cell.value != "":
                _apply_summary_color(
                    ws, dr, ci, val, is_pp=(lbl in pp_metrics))
            else:
                cell.fill = PatternFill("solid", fgColor=bg)
                cell.font = Font(name="Arial", size=9)

        ws.row_dimensions[dr].height = 20

    # Delta vs benchmark section
    dr = 4 + len(df_sum_sorted) + 1
    ws.merge_cells(f"A{dr}:{last_col}{dr}")
    hdr = ws[f"A{dr}"]
    hdr.value = f"▶  Delta vs Broad Market Avg (Sector Avg − Market Avg)"
    hdr.font = Font(bold=True, color=WHITE, name="Arial", size=10)
    hdr.fill = PatternFill("solid", fgColor="2E75B6")
    hdr.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[dr].height = 22
    dr += 1

    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=dr, column=ci, value=h)
        cell.font = Font(bold=True, color=WHITE, name="Arial", size=9)
        cell.fill = PatternFill("solid", fgColor="2E75B6")
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True)
        cell.border = _border("FFFFFF")
    ws.row_dimensions[dr].height = 40
    dr += 1

    for ri, (_, rd) in enumerate(df_sum_sorted.iterrows()):
        even = (ri % 2 == 0)
        bg = LGRAY if even else WHITE

        sc = ws.cell(row=dr, column=1, value=rd["Sector"])
        sc.font = Font(name="Arial", size=9, bold=True)
        sc.fill = PatternFill("solid", fgColor=bg)
        sc.alignment = Alignment(vertical="center")
        sc.border = _border()

        cc = ws.cell(row=dr, column=2, value=rd["Companies"])
        cc.font = Font(name="Arial", size=9)
        cc.fill = PatternFill("solid", fgColor=bg)
        cc.alignment = Alignment(horizontal="center", vertical="center")
        cc.border = _border()

        for ci, lbl in enumerate(metric_labels, 3):
            sec_val = rd.get(lbl)
            bench_val = n500_avgs.get(lbl)
            cell = ws.cell(row=dr, column=ci)
            if (sec_val is not None and not (isinstance(sec_val, float) and pd.isna(sec_val))
                    and bench_val is not None):
                delta = round(sec_val - bench_val, 1)
                cell.value = delta
                cell.number_format = '+0.0;-0.0;0.0'
            else:
                cell.value = ""
                delta = None
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = _border()
            if delta is not None:
                _apply_summary_color(
                    ws, dr, ci, delta, is_pp=(lbl in pp_metrics))
            else:
                cell.fill = PatternFill("solid", fgColor=bg)
                cell.font = Font(name="Arial", size=9)

        ws.row_dimensions[dr].height = 20
        dr += 1

    # Column widths
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 12
    for ci in range(3, TOTAL_COLS + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 15

    ws.freeze_panes = "A4"

    # Legend
    legend_row = dr + 1
    ws.merge_cells(f"A{legend_row}:{last_col}{legend_row}")
    leg = ws[f"A{legend_row}"]
    leg.value = ("Growth: >=15% = dark green | 5–15% = light green | 0–5% = amber | −10–0% = light red | <−10% = dark red  ||  "
                 "Margin pp: >=+2pp = dark green | 0–+2pp = light green | −2–0pp = light red | <−2pp = dark red")
    leg.font = Font(name="Arial", size=8, italic=True, color="595959")
    leg.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[legend_row].height = 18

    print(
        f"  [OK] Sector Summary sheet: {len(sectors)} sectors, {len(df_all)} companies")
    return df_sum, n500_avgs


# ══════════════════════════════════════════════════════════════════════════
# HTML INTERACTIVE DASHBOARD
# ══════════════════════════════════════════════════════════════════════════

def build_html_dashboard(df_sum, n500_avgs, output_path):
    if df_sum is None or df_sum.empty or "Sector" not in df_sum.columns:
        print("  Skipping HTML dashboard - no sector summary data available yet.")
        return
    metrics = [m[0] for m in SUMMARY_METRICS]
    pp_metrics_set = {m[0] for m in SUMMARY_METRICS if "pp" in m[0].lower()}
    sectors = df_sum["Sector"].tolist()

    sector_js = json.dumps(sectors)
    metrics_js = json.dumps(metrics)
    pp_metrics_js = json.dumps(list(pp_metrics_set))
    benchmark_js = json.dumps({k: (round(float(v), 1) if v is not None else None)
                               for k, v in n500_avgs.items()})

    data_by_metric = {}
    for m in metrics:
        data_by_metric[m] = []
        for _, row in df_sum.iterrows():
            v = row.get(m)
            data_by_metric[m].append(
                round(float(v), 1) if (v is not None and not (
                    isinstance(v, float) and pd.isna(v))) else None
            )
    data_js = json.dumps(data_by_metric)

    ts = datetime.today().strftime('%d %b %Y, %H:%M')

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Sector Dashboard · MCap >= {MCAP_FLOOR} Cr</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"></script>
<style>
  :root {{
    --navy: #0E2841; --blue: #243F8D; --light: #F2F2F2;
    --green: #00B050; --red: #C00000; --amber: #FFC000;
  }}
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{ font-family: Arial, sans-serif; background: #f5f7fa; color: #222; }}
  header {
    background: #0a0e14 !important;
    backdrop-filter: blur(12px);
    border-bottom: 1px solid rgba(255,255,255,0.1);
    padding: 0.75rem 2%;
    position: sticky;
    top: 0;
    z-index: 1000;
    display: grid;
    grid-template-columns: 320px 1fr 320px;
    align-items: center;
    color: #fff;
  }
  .header-left { justify-self: start; display: flex; align-items: center; gap: 0.8rem; }
  .logo { font-size: 0.95rem; font-weight: 700; color: #fff; text-transform: uppercase; letter-spacing: 1px; }
  .logo span { color: #3b82f6; }
  .nav-links { justify-self: center; display: flex; gap: 0.3rem; }
  .nav-links a { 
      color: #94a3b8; 
      text-decoration: none; 
      font-weight: 600; 
      font-size: 0.82rem; 
      transition: 0.3s; 
      padding: 0.4rem 0.8rem; 
      border-radius: 8px; 
      display: inline-block;
  }
  .nav-links a:hover, .nav-links a.active { color: #fff; background: rgba(255,255,255,0.08); }
  .nav-links a.active { color: #3b82f6 !important; background: rgba(255,255,255,0.06); }
  .header-right { justify-self: end; font-size: 0.75rem; color: #94a3b8; }
  .controls {{
    background: #fff; padding: 16px 28px; display: flex; gap: 16px;
    flex-wrap: wrap; align-items: center; border-bottom: 2px solid #e0e4ed;
  }}
  .controls label {{ font-size: .82rem; font-weight: bold; color: var(--navy); }}
  select, button {{
    padding: 7px 14px; border-radius: 6px; border: 1.5px solid #c0c8dc;
    font-size: .85rem; cursor: pointer; background: #fff;
  }}
  button.active {{ background: var(--blue); color: #fff; border-color: var(--blue); }}
  .chart-wrap {{
    background: #fff; border-radius: 10px; margin: 20px 28px;
    padding: 20px; box-shadow: 0 2px 8px rgba(0,0,0,.07);
  }}
  .chart-wrap h2 {{ font-size: 1rem; margin-bottom: 14px; color: var(--navy); }}
  canvas {{ max-height: 520px; }}
  .table-wrap {{
    background: #fff; border-radius: 10px; margin: 0 28px 28px;
    padding: 20px; box-shadow: 0 2px 8px rgba(0,0,0,.07); overflow-x: auto;
  }}
  table {{ width: 100%; border-collapse: collapse; font-size: .78rem; min-width: 900px; }}
  th {{
    background: var(--blue); color: #fff; padding: 8px 10px;
    text-align: center; white-space: nowrap; position: sticky; top: 0;
  }}
  th:first-child {{ text-align: left; min-width: 160px; }}
  td {{ padding: 6px 10px; border-bottom: 1px solid #eee; text-align: center; }}
  td:first-child {{ text-align: left; font-weight: bold; }}
  tr:hover td {{ background: #f0f4ff; }}
  .benchmark-row td {{
    background: #1F4E79 !important; color: #fff !important;
    font-weight: bold; font-size: .82rem;
  }}
  .badge {{
    display: inline-block; padding: 2px 7px; border-radius: 4px;
    font-size: .75rem; font-weight: bold;
  }}
  .g2 {{ background:#00B050;color:#fff; }}
  .g1 {{ background:#92D050;color:#222; }}
  .a0 {{ background:#FFEB9C;color:#555; }}
  .r1 {{ background:#FFC7CE;color:#333; }}
  .r2 {{ background:#C00000;color:#fff; }}
  .pp-g2 {{ background:#00B050;color:#fff; }}
  .pp-g1 {{ background:#92D050;color:#222; }}
  .pp-r1 {{ background:#FFC7CE;color:#333; }}
  .pp-r2 {{ background:#C00000;color:#fff; }}
  .na {{ color:#aaa; font-style:italic; }}
  .delta-pos {{ color: var(--green); font-weight: bold; }}
  .delta-neg {{ color: var(--red); font-weight: bold; }}
  .col-group-growth {{ border-left: 2px solid #243F8D; }}
  .col-group-margin {{ border-left: 2px solid #00B050; }}
  .col-group-annual {{ border-left: 2px solid #FF6600; }}
  footer {{
    text-align: center; padding: 14px; font-size: .78rem;
    color: #888; border-top: 1px solid #e0e4ed; margin-top: 8px;
  }}
</style>
</head>
<body>
<header>
  <div class="header-left">
    <div class="logo">Ashika <span>Results</span></div>
  </div>
  <nav class="nav-links">
    <a href="volume_dashboard.html">Quadrants</a>
    <a href="analytics.html">Analytics</a>
    <a href="results_dashboard.html">Results</a>
    <a href="results_dashboard.html#board">Board</a>
  </nav>
  <div class="header-right">
    <span id="ts">Generated: {ts}</span>
  </div>
</header>

<div class="controls">
  <div>
    <label>Metric&nbsp;</label>
    <select id="metricSel"></select>
  </div>
  <div>
    <label>View&nbsp;</label>
    <button class="active" id="btnAbsolute" onclick="setView('absolute')">Absolute Avg</button>
    <button id="btnDelta" onclick="setView('delta')">vs Benchmark Δ</button>
  </div>
  <div>
    <label>Sort&nbsp;</label>
    <button class="active" id="btnDesc" onclick="setSort('desc')">High -> Low</button>
    <button id="btnAsc" onclick="setSort('asc')">Low -> High</button>
    <button id="btnAlpha" onclick="setSort('alpha')">A–Z</button>
  </div>
  <div>
    <label>Top&nbsp;</label>
    <select id="topN">
      <option value="0">All sectors</option>
      <option value="10">Top 10</option>
      <option value="15">Top 15</option>
      <option value="20">Top 20</option>
    </select>
  </div>
</div>

<div class="chart-wrap">
  <h2 id="chartTitle">Loading...</h2>
  <canvas id="barChart"></canvas>
</div>

<div class="table-wrap">
  <table id="summaryTable">
    <thead>
      <tr id="tableHeadTop"></tr>
      <tr id="tableHead"></tr>
    </thead>
    <tbody id="tableBody"></tbody>
  </table>
</div>

<footer>Sector Analysis Dashboard · Growth rates in % · Margin changes in percentage points (pp) · Benchmark = broad market avg</footer>

<script>
const SECTORS    = {sector_js};
const METRICS    = {metrics_js};
const PP_METRICS = new Set({pp_metrics_js});
const BENCHMARK  = {benchmark_js};
const DATA       = {data_js};

// Group metrics for column headers
const COL_GROUPS = [
  {{ label: "Quarterly Growth (YoY)", cols: ["Avg Sales YoY Q%","Avg EBITDA YoY Q%","Avg NP YoY Q%","Avg EPS YoY Q%"], cls: "col-group-growth" }},
  {{ label: "Quarterly Margins", cols: ["Avg EBITDA Margin%","Avg PAT Margin%"], cls: "col-group-margin" }},
  {{ label: "Quarterly Margin Changes", cols: ["Avg EBITDA Margin YoY pp","Avg PAT Margin YoY pp","Avg EBITDA Margin QoQ pp","Avg PAT Margin QoQ pp"], cls: "col-group-margin" }},
  {{ label: "Annual Growth (FY YoY)", cols: ["Avg FY Sales YoY%","Avg FY EBITDA YoY%","Avg FY NP YoY%","Avg FY EPS YoY%"], cls: "col-group-annual" }},
  {{ label: "Annual Margins & Changes", cols: ["Avg FY EBITDA Margin%","Avg FY PAT Margin%","Avg FY EBITDA Margin YoY pp","Avg FY PAT Margin YoY pp"], cls: "col-group-annual" }},
];

let currentMetric = METRICS[0];
let currentView   = 'absolute';
let currentSort   = 'desc';
let topN          = 0;
let chart         = null;

function val(sector, metric) {{
  const idx = SECTORS.indexOf(sector);
  const arr = DATA[metric];
  return (arr && idx >= 0) ? arr[idx] : null;
}}

function badge(v, isPP) {{
  if (v === null || v === undefined) return '<span class="na">-</span>';
  const n = parseFloat(v);
  let cls;
  if (isPP) {{
    cls = n >= 2 ? 'pp-g2' : n >= 0 ? 'pp-g1' : n >= -2 ? 'pp-r1' : 'pp-r2';
    const sign = n >= 0 ? '+' : '';
    return `<span class="badge ${{cls}}">${{sign}}${{n.toFixed(1)}}pp</span>`;
  }} else {{
    cls = n >= 15 ? 'g2' : n >= 5 ? 'g1' : n >= 0 ? 'a0' : n >= -10 ? 'r1' : 'r2';
    return `<span class="badge ${{cls}}">${{n.toFixed(1)}}%</span>`;
  }}
}}

function deltaHtml(d, isPP) {{
  if (d === null || d === undefined) return '<span class="na">-</span>';
  const n = parseFloat(d);
  const cls = n >= 0 ? 'delta-pos' : 'delta-neg';
  const sign = n >= 0 ? '+' : '';
  const unit = isPP ? 'pp' : '';
  return `<span class="${{cls}}">${{sign}}${{n.toFixed(1)}}${{unit}}</span>`;
}}

function barColor(v) {{
  if (v === null) return '#ccc';
  const isPP = PP_METRICS.has(currentMetric);
  if (isPP) {{
    return v >= 2 ? '#00B050' : v >= 0 ? '#92D050' : v >= -2 ? '#FF9999' : '#C00000';
  }}
  return v >= 15 ? '#00B050' : v >= 5 ? '#92D050' : v >= 0 ? '#FFC000' : v >= -10 ? '#FF9999' : '#C00000';
}}

function getSortedSectors() {{
  let pairs = SECTORS.map(s => ({{
    sector: s,
    value: currentView === 'delta'
      ? (val(s, currentMetric) !== null && BENCHMARK[currentMetric] !== null
          ? parseFloat((val(s, currentMetric) - BENCHMARK[currentMetric]).toFixed(1))
          : null)
      : val(s, currentMetric)
  }}));
  if (currentSort === 'desc') {{
    pairs.sort((a, b) => {{ if (a.value === null) return 1; if (b.value === null) return -1; return b.value - a.value; }});
  }} else if (currentSort === 'asc') {{
    pairs.sort((a, b) => {{ if (a.value === null) return 1; if (b.value === null) return -1; return a.value - b.value; }});
  }} else {{
    pairs.sort((a, b) => a.sector.localeCompare(b.sector));
  }}
  if (topN > 0) pairs = pairs.slice(0, topN);
  return pairs;
}}

function renderChart() {{
  const pairs  = getSortedSectors();
  const labels = pairs.map(p => p.sector);
  const values = pairs.map(p => p.value);
  const colors = values.map(barColor);
  const bench  = BENCHMARK[currentMetric];
  const isPP   = PP_METRICS.has(currentMetric);
  const unit   = isPP ? 'pp' : '%';

  const datasets = [{{
    label: currentView === 'delta' ? 'Delta vs Benchmark' : (isPP ? 'Avg pp' : 'Avg %'),
    data: values,
    backgroundColor: colors,
    borderRadius: 4,
  }}];

  if (currentView === 'absolute' && bench !== null) {{
    datasets.push({{
      type: 'line',
      label: 'Broad Market Avg',
      data: Array(labels.length).fill(bench),
      borderColor: '#FF6600',
      borderWidth: 2,
      borderDash: [6, 3],
      pointRadius: 0,
      fill: false,
    }});
  }}

  const title = currentView === 'delta'
    ? `${{currentMetric}} - Delta vs Broad Market (${{bench !== null ? bench.toFixed(1)+unit : 'N/A'}})`
    : `${{currentMetric}} - Sector Averages`;
  document.getElementById('chartTitle').textContent = title;

  if (chart) chart.destroy();
  chart = new Chart(document.getElementById('barChart'), {{
    type: 'bar',
    data: {{ labels, datasets }},
    options: {{
      indexAxis: 'y',
      responsive: true,
      plugins: {{
        legend: {{ display: true, position: 'top' }},
        tooltip: {{
          callbacks: {{
            label: ctx => ` ${{ctx.parsed.x !== null ? ctx.parsed.x.toFixed(1) + unit : '-'}}`
          }}
        }}
      }},
      scales: {{
        x: {{ title: {{ display: true, text: isPP ? 'Percentage Points' : '% Growth' }}, grid: {{ color: '#eee' }} }},
        y: {{ ticks: {{ font: {{ size: 10 }} }} }}
      }}
    }}
  }});
}}

function renderTable() {{
  const headTop = document.getElementById('tableHeadTop');
  const head    = document.getElementById('tableHead');
  const body    = document.getElementById('tableBody');
  const bench   = BENCHMARK[currentMetric];

  // Build grouped header row
  let topHtml = '<th rowspan="2">Sector</th>';
  COL_GROUPS.forEach((g, gi) => {{
    topHtml += `<th colspan="${{g.cols.length}}" class="${{g.cls}}" style="border-bottom:1px solid rgba(255,255,255,0.3)">${{g.label}}</th>`;
  }});
  topHtml += '<th rowspan="2">vs Benchmark</th>';
  headTop.innerHTML = topHtml;

  let subHtml = '';
  COL_GROUPS.forEach((g, gi) => {{
    g.cols.forEach((c, ci) => {{
      const short = c.replace('Avg ', '').replace(' YoY Q%','').replace(' QoQ pp','').replace(' YoY pp','').replace(' YoY%','').replace(' Margin%','').replace(' Margin ','');
      subHtml += `<th class="${{gi>0 && ci===0 ? g.cls : ''}}">${{c.replace('Avg ','').replace(' Q%','%')}}</th>`;
    }});
  }});
  head.innerHTML = subHtml;

  // Benchmark row
  let brow = '<tr class="benchmark-row"><td>▶ Broad Market Avg</td>';
  METRICS.forEach(m => {{
    const bv = BENCHMARK[m];
    const isPP = PP_METRICS.has(m);
    const unit = isPP ? 'pp' : '%';
    brow += `<td>${{bv !== null ? (isPP && bv >= 0 ? '+' : '') + bv.toFixed(1) + unit : '-'}}</td>`;
  }});
  brow += '<td>-</td></tr>';

  // Data rows
  const pairs = getSortedSectors();
  let rows = '';
  pairs.forEach((p) => {{
    const s = p.sector;
    rows += `<tr><td>${{s}}</td>`;
    METRICS.forEach(m => {{
      const isPP = PP_METRICS.has(m);
      rows += `<td>${{badge(val(s, m), isPP)}}</td>`;
    }});
    const dv = (val(s, currentMetric) !== null && bench !== null)
      ? val(s, currentMetric) - bench : null;
    rows += `<td>${{deltaHtml(dv, PP_METRICS.has(currentMetric))}}</td>`;
    rows += '</tr>';
  }});

  body.innerHTML = brow + rows;
}}

function refresh() {{ renderChart(); renderTable(); }}

function setView(v) {{
  currentView = v;
  document.getElementById('btnAbsolute').className = v === 'absolute' ? 'active' : '';
  document.getElementById('btnDelta').className    = v === 'delta'    ? 'active' : '';
  refresh();
}}

function setSort(s) {{
  currentSort = s;
  ['btnDesc','btnAsc','btnAlpha'].forEach(id => document.getElementById(id).className = '');
  document.getElementById(s === 'desc' ? 'btnDesc' : s === 'asc' ? 'btnAsc' : 'btnAlpha').className = 'active';
  refresh();
}}

const sel = document.getElementById('metricSel');
// Group metrics in select
const groups = [
  {{ label: 'Quarterly Growth (YoY)', metrics: METRICS.slice(0, 4) }},
  {{ label: 'Quarterly Margins', metrics: METRICS.slice(4, 6) }},
  {{ label: 'Quarterly Margin Changes', metrics: METRICS.slice(6, 10) }},
  {{ label: 'Annual Growth (FY YoY)', metrics: METRICS.slice(10, 14) }},
  {{ label: 'Annual Margins & Changes', metrics: METRICS.slice(14) }},
];
groups.forEach(g => {{
  const og = document.createElement('optgroup');
  og.label = g.label;
  g.metrics.forEach(m => {{
    const o = document.createElement('option');
    o.value = m; o.textContent = m;
    og.appendChild(o);
  }});
  sel.appendChild(og);
}});
sel.addEventListener('change', e => {{ currentMetric = e.target.value; refresh(); }});
document.getElementById('topN').addEventListener('change', e => {{
  topN = parseInt(e.target.value); refresh();
}});

refresh();
</script>
</body>
</html>"""

    with open(output_path, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"  [OK] Dashboard saved: {output_path}")


# ══════════════════════════════════════════════════════════════════════════
# DAILY HTML DASHBOARD
# ══════════════════════════════════════════════════════════════════════════

def build_daily_dashboard(df_date, output_path, date_label=""):
    """
    Build an interactive HTML dashboard for a single day's (or date-range's) results.
    df_date: DataFrame with ALL_COLS columns, already MCap-filtered.
    date_label: human-readable string shown in the header (e.g. "14 Apr 2026" or "01–14 Apr 2026").
    """
    if df_date is None or df_date.empty:
        print("  Skipping Daily Dashboard - no data available.")
        return

    # Columns shown in the company table (skip raw prev-qtr/prev-fy reference cols)
    DISPLAY_COLS = [
        "Result Date", "Company Name", "Sector",
        "Quarter",
        "Sales (Cr)", "EBITDA (Cr)", "Net Profit (Cr)", "EPS (Rs)",
        "EBITDA Margin%", "PAT Margin%",
        "Sales YoY Q%", "EBITDA YoY Q%", "NP YoY Q%", "EPS YoY Q%",
        "EBITDA Margin YoY pp", "PAT Margin YoY pp",
        "Sales QoQ%", "EBITDA QoQ%", "NP QoQ%", "EPS QoQ%",
        "EBITDA Margin QoQ pp", "PAT Margin QoQ pp",
        "FY Sales YoY%", "FY EBITDA YoY%", "FY NP YoY%", "FY EPS YoY%",
        "FY EBITDA Margin YoY pp", "FY PAT Margin YoY pp",
    ]
    SORT_OPTIONS = [
        "Sales YoY Q%", "EBITDA YoY Q%", "NP YoY Q%", "EPS YoY Q%",
        "Sales QoQ%", "NP QoQ%",
        "FY Sales YoY%", "FY NP YoY%",
        "Company Name",
    ]
    PCT_DAILY = {
        "Sales YoY Q%", "EBITDA YoY Q%", "NP YoY Q%", "EPS YoY Q%",
        "Sales QoQ%", "EBITDA QoQ%", "NP QoQ%", "EPS QoQ%",
        "FY Sales YoY%", "FY EBITDA YoY%", "FY NP YoY%", "FY EPS YoY%",
    }
    PP_DAILY = {
        "EBITDA Margin YoY pp", "PAT Margin YoY pp",
        "EBITDA Margin QoQ pp", "PAT Margin QoQ pp",
        "FY EBITDA Margin YoY pp", "FY PAT Margin YoY pp",
    }

    # Ensure numeric cols are numeric
    for c in DISPLAY_COLS:
        if c in df_date.columns and c not in {"Result Date", "Company Name", "Sector", "Quarter"}:
            df_date[c] = pd.to_numeric(df_date[c], errors="coerce")

    # Build records for JS - only cols that exist
    existing = [c for c in DISPLAY_COLS if c in df_date.columns]
    records = []
    for _, row in df_date.iterrows():
        rec = {}
        for c in existing:
            v = row.get(c)
            if isinstance(v, pd.Timestamp):
                rec[c] = v.strftime("%d %b %Y")
            elif v is None or (isinstance(v, float) and pd.isna(v)):
                rec[c] = None
            else:
                rec[c] = v
        records.append(rec)

    # Available dates for the date-range filter
    date_vals = sorted(df_date["Result Date"].dropna().unique(
    ).tolist()) if "Result Date" in df_date.columns else []
    # Convert Timestamps to strings if needed
    date_vals = [d.strftime("%d %b %Y") if isinstance(
        d, pd.Timestamp) else str(d) for d in date_vals]

    sectors = sorted(df_date["Sector"].dropna().unique(
    ).tolist()) if "Sector" in df_date.columns else []

    data_js = json.dumps(records)
    cols_js = json.dumps(existing)
    sort_opts_js = json.dumps([c for c in SORT_OPTIONS if c in existing])
    pct_js = json.dumps(list(PCT_DAILY & set(existing)))
    pp_js = json.dumps(list(PP_DAILY & set(existing)))
    dates_js = json.dumps(date_vals)
    sectors_js = json.dumps(sectors)
    ts = datetime.today().strftime("%d %b %Y, %H:%M")

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Daily Results Dashboard · {date_label}</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"></script>
<style>
  :root {{
    --navy:#0E2841; --blue:#243F8D; --light:#F2F2F2;
    --green:#00B050; --red:#C00000; --amber:#FFC000;
  }}
  *{{box-sizing:border-box;margin:0;padding:0}}
  body{{font-family:Arial,sans-serif;background:#f5f7fa;color:#222}}
  header {
    background: #0a0e14 !important;
    backdrop-filter: blur(12px);
    border-bottom: 1px solid rgba(255,255,255,0.1);
    padding: 0.75rem 2%;
    position: sticky;
    top: 0;
    z-index: 1000;
    display: grid;
    grid-template-columns: 320px 1fr 320px;
    align-items: center;
    color: #fff;
  }
  .header-left { justify-self: start; display: flex; align-items: center; gap: 0.8rem; }
  .logo { font-size: 0.95rem; font-weight: 700; color: #fff; text-transform: uppercase; letter-spacing: 1px; }
  .logo span { color: #3b82f6; }
  .nav-links { justify-self: center; display: flex; gap: 0.3rem; }
  .nav-links a { 
      color: #94a3b8; 
      text-decoration: none; 
      font-weight: 600; 
      font-size: 0.82rem; 
      transition: 0.3s; 
      padding: 0.4rem 0.8rem; 
      border-radius: 8px; 
      display: inline-block;
  }
  .nav-links a:hover, .nav-links a.active { color: #fff; background: rgba(255,255,255,0.08); }
  .nav-links a.active { color: #3b82f6 !important; background: rgba(255,255,255,0.06); }
  .header-right { justify-self: end; font-size: 0.75rem; color: #94a3b8; }
  .controls{{background:#fff;padding:12px 24px;display:flex;gap:12px;flex-wrap:wrap;align-items:flex-end;border-bottom:2px solid #e0e4ed}}
  .ctrl-group{{display:flex;flex-direction:column;gap:4px}}
  .ctrl-group label{{font-size:.75rem;font-weight:bold;color:var(--navy)}}
  select,input[type=text]{{padding:6px 10px;border-radius:5px;border:1.5px solid #c0c8dc;font-size:.82rem;background:#fff;cursor:pointer}}
  input[type=text]{{width:140px}}
  .btn-row{{display:flex;gap:6px;align-items:center;padding-top:2px}}
  button{{padding:6px 12px;border-radius:5px;border:1.5px solid #c0c8dc;font-size:.82rem;cursor:pointer;background:#fff}}
  button.active{{background:var(--blue);color:#fff;border-color:var(--blue)}}
  .stats-bar{{background:#fff;padding:10px 24px;display:flex;gap:20px;flex-wrap:wrap;border-bottom:1px solid #e0e4ed;font-size:.82rem}}
  .stat{{display:flex;flex-direction:column}}
  .stat .sv{{font-size:1.1rem;font-weight:bold;color:var(--navy)}}
  .stat .sl{{color:#888;font-size:.72rem}}
  .chart-wrap{{background:#fff;border-radius:10px;margin:16px 24px;padding:18px;box-shadow:0 2px 8px rgba(0,0,0,.07)}}
  .chart-wrap h2{{font-size:.95rem;margin-bottom:12px;color:var(--navy)}}
  canvas{{max-height:420px}}
  .table-wrap{{background:#fff;border-radius:10px;margin:0 24px 24px;padding:16px;box-shadow:0 2px 8px rgba(0,0,0,.07);overflow-x:auto}}
  table{{width:100%;border-collapse:collapse;font-size:.75rem;min-width:1100px}}
  th{{background:var(--blue);color:#fff;padding:7px 8px;text-align:center;white-space:nowrap;position:sticky;top:0;z-index:2;cursor:pointer}}
  th:hover{{background:#1a2e6e}}
  th.sort-asc::after{{content:" ▲"}}
  th.sort-desc::after{{content:" ▼"}}
  th:first-child,th:nth-child(2),th:nth-child(3){{text-align:left;position:sticky;z-index:3}}
  td{{padding:5px 8px;border-bottom:1px solid #eee;text-align:center;white-space:nowrap}}
  td:first-child,td:nth-child(2){{text-align:left}}
  td:nth-child(3){{text-align:left;font-size:.72rem;color:#555}}
  tr:hover td{{background:#f0f4ff!important}}
  .badge{{display:inline-block;padding:2px 6px;border-radius:4px;font-size:.72rem;font-weight:bold}}
  .g2{{background:#00B050;color:#fff}} .g1{{background:#92D050;color:#222}}
  .a0{{background:#FFEB9C;color:#555}} .r1{{background:#FFC7CE;color:#333}}
  .r2{{background:#C00000;color:#fff}} .na{{color:#bbb;font-style:italic}}
  .pp-g2{{background:#00B050;color:#fff}} .pp-g1{{background:#92D050;color:#222}}
  .pp-r1{{background:#FFC7CE;color:#333}} .pp-r2{{background:#C00000;color:#fff}}
  .mcap{{font-size:.72rem;color:#555}}
  .grp-hdr th{{background:#1F4E79!important;font-size:.7rem;border-right:1px solid rgba(255,255,255,.2)}}
  .date-range-row{{display:flex;gap:8px;align-items:center;flex-wrap:wrap}}
  .date-range-row label{{font-size:.75rem;font-weight:bold;color:var(--navy)}}
  .date-chip{{display:inline-block;padding:3px 10px;border-radius:12px;background:#e8ecf8;font-size:.75rem;cursor:pointer;border:1.5px solid transparent}}
  .date-chip.active{{background:var(--blue);color:#fff;border-color:var(--blue)}}
  footer{{text-align:center;padding:12px;font-size:.75rem;color:#888;border-top:1px solid #e0e4ed}}
  #noData{{display:none;padding:40px;text-align:center;color:#888;font-size:1rem}}
</style>
</head>
<body>
<header>
  <div class="header-left">
    <div class="logo">Ashika <span>Results</span></div>
  </div>
  <nav class="nav-links">
    <a href="volume_dashboard.html">Quadrants</a>
    <a href="analytics.html">Analytics</a>
    <a href="results_dashboard.html">Results</a>
    <a href="results_dashboard.html#board">Board</a>
  </nav>
  <div class="header-right">
    <span id="headerDate">{date_label} &nbsp;·&nbsp; Generated: {ts}</span>
  </div>
</header>

<div class="controls">
  <div class="ctrl-group">
    <label>Date Range</label>
    <div class="date-range-row">
      <label>From</label>
      <select id="dateFrom"></select>
      <label>To</label>
      <select id="dateTo"></select>
      <button onclick="applyDateRange()">Apply</button>
      <button onclick="resetDateRange()">All Dates</button>
    </div>
  </div>
  <div class="ctrl-group">
    <label>Sector Filter</label>
    <select id="sectorFilter">
      <option value="">All Sectors</option>
    </select>
  </div>
  <div class="ctrl-group">
    <label>Sort By</label>
    <select id="sortCol"></select>
  </div>
  <div class="ctrl-group">
    <label>Order</label>
    <div class="btn-row">
      <button class="active" id="btnDesc" onclick="setOrder('desc')">High -> Low</button>
      <button id="btnAsc" onclick="setOrder('asc')">Low -> High</button>
    </div>
  </div>
  <div class="ctrl-group">
    <label>Chart Metric</label>
    <select id="chartMetricSel"></select>
  </div>
  <div class="ctrl-group">
    <label>Search Company</label>
    <input type="text" id="searchBox" placeholder="Type to filter..." oninput="refresh()">
  </div>
</div>

<div class="stats-bar" id="statsBar"></div>

<div class="chart-wrap">
  <h2 id="chartTitle">Loading...</h2>
  <canvas id="barChart"></canvas>
</div>

<div class="table-wrap">
  <div id="noData">No companies match the current filters.</div>
  <table id="mainTable">
    <thead>
      <tr class="grp-hdr" id="groupHdr"></tr>
      <tr id="colHdr"></tr>
    </thead>
    <tbody id="tableBody"></tbody>
  </table>
</div>

<footer>MCap >= {MCAP_FLOOR} Cr · YoY/QoQ growth in % · Margin changes in pp · Click column headers to sort</footer>

<script>
const ALL_DATA    = {data_js};
const COLS        = {cols_js};
const SORT_OPTS   = {sort_opts_js};
const PCT_COLS    = new Set({pct_js});
const PP_COLS     = new Set({pp_js});
const ALL_DATES   = {dates_js};
const ALL_SECTORS = {sectors_js};

// Column groups for the header
const COL_GROUPS = [
  {{ label: "Identity",        cols: ["Result Date","Company Name","Sector","Market Cap (Cr)","Quarter"] }},
  {{ label: "Quarterly Financials", cols: ["Sales (Cr)","EBITDA (Cr)","Net Profit (Cr)","EPS (Rs)","EBITDA Margin%","PAT Margin%"] }},
  {{ label: "Quarterly YoY Growth", cols: ["Sales YoY Q%","EBITDA YoY Q%","NP YoY Q%","EPS YoY Q%","EBITDA Margin YoY pp","PAT Margin YoY pp"] }},
  {{ label: "Quarterly QoQ Growth", cols: ["Sales QoQ%","EBITDA QoQ%","NP QoQ%","EPS QoQ%","EBITDA Margin QoQ pp","PAT Margin QoQ pp"] }},
  {{ label: "Annual FY YoY",   cols: ["FY Sales YoY%","FY EBITDA YoY%","FY NP YoY%","FY EPS YoY%","FY EBITDA Margin YoY pp","FY PAT Margin YoY pp"] }},
];

const CHART_METRICS = [
  "Sales YoY Q%","EBITDA YoY Q%","NP YoY Q%","EPS YoY Q%",
  "Sales QoQ%","NP QoQ%","FY Sales YoY%","FY NP YoY%",
].filter(m => COLS.includes(m));

let sortCol   = SORT_OPTS[0] || "Market Cap (Cr)";
let sortOrder = 'desc';
let dateFrom  = ALL_DATES.length ? ALL_DATES[0] : null;
let dateTo    = ALL_DATES.length ? ALL_DATES[ALL_DATES.length-1] : null;
let chart     = null;

// -- Populate controls -------------------------------------------------
const dfSel = document.getElementById('dateFrom');
const dtSel = document.getElementById('dateTo');
ALL_DATES.forEach(d => {{
  dfSel.innerHTML += `<option value="${{d}}">${{d}}</option>`;
  dtSel.innerHTML += `<option value="${{d}}">${{d}}</option>`;
}});
if (ALL_DATES.length) {{
  dfSel.value = ALL_DATES[0];
  dtSel.value = ALL_DATES[ALL_DATES.length-1];
}}

const secSel = document.getElementById('sectorFilter');
ALL_SECTORS.forEach(s => secSel.innerHTML += `<option value="${{s}}">${{s}}</option>`);

const sortSel = document.getElementById('sortCol');
SORT_OPTS.forEach(c => sortSel.innerHTML += `<option value="${{c}}">${{c}}</option>`);
sortSel.addEventListener('change', e => {{ sortCol = e.target.value; refresh(); }});
secSel.addEventListener('change', () => refresh());

const chartSel = document.getElementById('chartMetricSel');
CHART_METRICS.forEach(m => chartSel.innerHTML += `<option value="${{m}}">${{m}}</option>`);
chartSel.addEventListener('change', () => renderChart(getFiltered()));

function applyDateRange() {{
  dateFrom = dfSel.value;
  dateTo   = dtSel.value;
  refresh();
}}
function resetDateRange() {{
  dateFrom = ALL_DATES.length ? ALL_DATES[0] : null;
  dateTo   = ALL_DATES.length ? ALL_DATES[ALL_DATES.length-1] : null;
  if (ALL_DATES.length) {{ dfSel.value = ALL_DATES[0]; dtSel.value = ALL_DATES[ALL_DATES.length-1]; }}
  refresh();
}}
function setOrder(o) {{
  sortOrder = o;
  document.getElementById('btnDesc').className = o==='desc'?'active':'';
  document.getElementById('btnAsc').className  = o==='asc' ?'active':'';
  refresh();
}}

// -- Filtering & sorting -----------------------------------------------
function dateIndex(d) {{ return ALL_DATES.indexOf(d); }}

function getFiltered() {{
  const sector = secSel.value;
  const q      = document.getElementById('searchBox').value.toLowerCase();
  let rows = ALL_DATA.filter(r => {{
    if (dateFrom && dateTo) {{
      const di = dateIndex(r["Result Date"]);
      const fi = dateIndex(dateFrom);
      const ti = dateIndex(dateTo);
      if (di < fi || di > ti) return false;
    }}
    if (sector && r["Sector"] !== sector) return false;
    if (q && !(r["Company Name"]||'').toLowerCase().includes(q)) return false;
    return true;
  }});

  rows.sort((a, b) => {{
    const av = a[sortCol], bv = b[sortCol];
    if (av === null && bv === null) return 0;
    if (av === null) return 1;
    if (bv === null) return -1;
    if (typeof av === 'string') return sortOrder==='asc' ? av.localeCompare(bv) : bv.localeCompare(av);
    return sortOrder==='asc' ? av - bv : bv - av;
  }});
  return rows;
}}

// -- Badge helpers -----------------------------------------------------
function badge(v, col) {{
  if (v === null || v === undefined) return '<span class="na">-</span>';
  const n = parseFloat(v);
  if (isNaN(n)) return `<span style="font-size:.72rem">${{v}}</span>`;
  const isPP = PP_COLS.has(col);
  const isPct = PCT_COLS.has(col);
  if (!isPP && !isPct) return n.toLocaleString('en-IN', {{maximumFractionDigits:0}});
  let cls;
  if (isPP) {{
    cls = n>=2?'pp-g2':n>=0?'pp-g1':n>=-2?'pp-r1':'pp-r2';
    return `<span class="badge ${{cls}}">${{n>=0?'+':''}}${{n.toFixed(1)}}pp</span>`;
  }} else {{
    cls = n>=15?'g2':n>=5?'g1':n>=0?'a0':n>=-10?'r1':'r2';
    return `<span class="badge ${{cls}}">${{n.toFixed(1)}}%</span>`;
  }}
}}

function barColor(v, metric) {{
  if (v===null) return '#ccc';
  const isPP = PP_COLS.has(metric);
  if (isPP) return v>=2?'#00B050':v>=0?'#92D050':v>=-2?'#FF9999':'#C00000';
  return v>=15?'#00B050':v>=5?'#92D050':v>=0?'#FFC000':v>=-10?'#FF9999':'#C00000';
}}

// -- Stats bar ---------------------------------------------------------
function renderStats(rows) {{
  const n = rows.length;
  const sectors = new Set(rows.map(r=>r["Sector"]).filter(Boolean)).size;
  const positiveYoY = rows.filter(r=>r["Sales YoY Q%"]!==null && r["Sales YoY Q%"]>=0).length;
  const negYoY = rows.filter(r=>r["Sales YoY Q%"]!==null && r["Sales YoY Q%"]<0).length;
  const avgSalesYoY = rows.filter(r=>r["Sales YoY Q%"]!==null).reduce((s,r)=>s+r["Sales YoY Q%"],0) / (rows.filter(r=>r["Sales YoY Q%"]!==null).length||1);
  const avgNPYoY    = rows.filter(r=>r["NP YoY Q%"]!==null).reduce((s,r)=>s+r["NP YoY Q%"],0) / (rows.filter(r=>r["NP YoY Q%"]!==null).length||1);
  document.getElementById('statsBar').innerHTML = `
    <div class="stat"><span class="sv">${{n}}</span><span class="sl">Companies</span></div>
    <div class="stat"><span class="sv">${{sectors}}</span><span class="sl">Sectors</span></div>
    <div class="stat"><span class="sv" style="color:var(--green)">${{positiveYoY}}</span><span class="sl">Sales YoY >= 0</span></div>
    <div class="stat"><span class="sv" style="color:var(--red)">${{negYoY}}</span><span class="sl">Sales YoY < 0</span></div>
    <div class="stat"><span class="sv">${{isFinite(avgSalesYoY)?avgSalesYoY.toFixed(1)+'%':'-'}}</span><span class="sl">Avg Sales YoY</span></div>
    <div class="stat"><span class="sv">${{isFinite(avgNPYoY)?avgNPYoY.toFixed(1)+'%':'-'}}</span><span class="sl">Avg NP YoY</span></div>
  `;
}}

// -- Chart -------------------------------------------------------------
function renderChart(rows) {{
  const metric = chartSel.value;
  if (!metric) return;

  // Group by sector, compute avg
  const secMap = {{}};
  rows.forEach(r => {{
    const s = r["Sector"] || "Unknown";
    if (!secMap[s]) secMap[s] = [];
    if (r[metric] !== null) secMap[s].push(r[metric]);
  }});
  let pairs = Object.entries(secMap)
    .map(([s, vals]) => ({{ s, v: vals.length ? vals.reduce((a,b)=>a+b,0)/vals.length : null }}))
    .filter(p => p.v !== null)
    .sort((a,b) => b.v - a.v);

  document.getElementById('chartTitle').textContent = `${{metric}} - Avg by Sector (${{rows.length}} companies)`;
  if (chart) chart.destroy();
  chart = new Chart(document.getElementById('barChart'), {{
    type: 'bar',
    data: {{
      labels: pairs.map(p=>p.s),
      datasets: [{{ label: metric, data: pairs.map(p=>parseFloat(p.v.toFixed(1))),
        backgroundColor: pairs.map(p=>barColor(p.v, metric)), borderRadius: 4 }}]
    }},
    options: {{
      indexAxis: 'y', responsive: true,
      plugins: {{ legend: {{ display: false }},
        tooltip: {{ callbacks: {{ label: ctx => ` ${{ctx.parsed.x !== null ? ctx.parsed.x.toFixed(1)+(PP_COLS.has(metric)?'pp':'%') : '-'}}` }} }} }},
      scales: {{
        x: {{ title: {{ display: true, text: PP_COLS.has(metric)?'pp':'%' }}, grid: {{ color:'#eee' }} }},
        y: {{ ticks: {{ font: {{ size: 10 }} }} }}
      }}
    }}
  }});
}}

// -- Table -------------------------------------------------------------
let tableSortCol = null;
let tableSortDir = 'desc';

function buildHeaders() {{
  const existCols = COLS;
  const grpHdr = document.getElementById('groupHdr');
  const colHdr = document.getElementById('colHdr');
  let topHtml = '', subHtml = '';
  COL_GROUPS.forEach(g => {{
    const visible = g.cols.filter(c => existCols.includes(c));
    if (!visible.length) return;
    topHtml += `<th colspan="${{visible.length}}" style="text-align:center">${{g.label}}</th>`;
    visible.forEach(c => {{
      const short = c.replace(' YoY Q%','').replace(' QoQ%','').replace(' YoY%','').replace(' YoY pp','').replace(' QoQ pp','').replace(' (Cr)','').replace(' (Rs)','').replace('Net Profit','NP').replace('Market Cap','MCap').replace('Result ','');
      subHtml += `<th onclick="tableSort('${{c}}')" id="th_${{c.replace(/[^a-zA-Z0-9]/g,'_')}}">${{short}}</th>`;
    }});
  }});
  grpHdr.innerHTML = topHtml;
  colHdr.innerHTML = subHtml;
}}

function tableSort(col) {{
  if (tableSortCol === col) {{ tableSortDir = tableSortDir==='desc'?'asc':'desc'; }}
  else {{ tableSortCol = col; tableSortDir = 'desc'; }}
  // Update header indicators
  document.querySelectorAll('#colHdr th').forEach(th => th.classList.remove('sort-asc','sort-desc'));
  const thId = 'th_' + col.replace(/[^a-zA-Z0-9]/g,'_');
  const thEl = document.getElementById(thId);
  if (thEl) thEl.classList.add(tableSortDir==='desc'?'sort-desc':'sort-asc');
  renderTable(getFiltered());
}}

function renderTable(rows) {{
  // Apply table-level sort if set
  if (tableSortCol) {{
    rows = [...rows].sort((a,b) => {{
      const av = a[tableSortCol], bv = b[tableSortCol];
      if (av===null&&bv===null) return 0;
      if (av===null) return 1; if (bv===null) return -1;
      if (typeof av==='string') return tableSortDir==='asc'?av.localeCompare(bv):bv.localeCompare(av);
      return tableSortDir==='asc'?av-bv:bv-av;
    }});
  }}

  const existCols = COLS;
  const noData = document.getElementById('noData');
  const tbl    = document.getElementById('mainTable');
  if (!rows.length) {{ noData.style.display='block'; tbl.style.display='none'; return; }}
  noData.style.display='none'; tbl.style.display='';

  let html = '';
  rows.forEach((r, ri) => {{
    const bg = ri%2===0 ? '#fff' : '#f8f9fd';
    html += `<tr style="background:${{bg}}">`;
    COL_GROUPS.forEach(g => {{
      g.cols.filter(c=>existCols.includes(c)).forEach(c => {{
        html += `<td>${{badge(r[c], c)}}</td>`;
      }});
    }});
    html += '</tr>';
  }});
  document.getElementById('tableBody').innerHTML = html;
}}

function refresh() {{
  const rows = getFiltered();
  renderStats(rows);
  renderChart(rows);
  renderTable(rows);
}}

buildHeaders();
refresh();
</script>
</body>
</html>"""

    with open(output_path, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"  [OK] Daily Dashboard saved: {output_path}")


# ══════════════════════════════════════════════════════════════════════════
# LOAD ALL SECTOR DATA FROM WORKBOOK
# ══════════════════════════════════════════════════════════════════════════

def load_all_sector_data(wb_s):
    frames = []
    for sname in wb_s.sheetnames:
        if sname == "Sector Summary":
            continue
        ws = wb_s[sname]
        df = read_sheet_df(ws)
        if not df.empty:
            frames.append(df)
    if frames:
        return pd.concat(frames, ignore_index=True)
    return pd.DataFrame(columns=ALL_COLS)


# ══════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════

def main():
    # -- Parse arguments ------------------------------------------------
    # Usage:
    #   py sorter.py                        -> yesterday
    #   py sorter.py 2026-04-14             -> single date
    #   py sorter.py 2026-04-01 2026-04-14  -> date range (inclusive)
    if len(sys.argv) >= 3:
        start_date = datetime.strptime(sys.argv[1], "%Y-%m-%d")
        end_date = datetime.strptime(sys.argv[2], "%Y-%m-%d")
    elif len(sys.argv) == 2:
        start_date = end_date = datetime.strptime(sys.argv[1], "%Y-%m-%d")
    else:
        start_date = end_date = datetime.today() - timedelta(days=1)

    # Build list of dates to process
    dates = []
    d = start_date
    while d <= end_date:
        dates.append(d)
        d += timedelta(days=1)

    if len(dates) == 1:
        date_label = dates[0].strftime("%d %b %Y")
    else:
        date_label = f"{dates[0].strftime('%d %b %Y')} – {dates[-1].strftime('%d %b %Y')}"

    print(f"\nResults Tracker - {date_label}  [MCap >= {MCAP_FLOOR} Cr filter]")
    if len(dates) > 1:
        print(
            f"  Processing {len(dates)} dates: {dates[0].strftime('%Y-%m-%d')} -> {dates[-1].strftime('%Y-%m-%d')}")
    print("=" * 60)

    # 1. Cache
    print("[1/4] Loading sector cache...")
    cache = load_cache()
    cache = build_bm_lookup(cache)
    save_cache(cache)

    # 2. Screener login
    print("[2/4] Logging into Screener...")
    session = screener_login()

    num_cols = [c for c in ALL_COLS if c not in {
        "Result Date", "Company Name", "Sector", "Quarter",
        "Screener Ticker", "Industry Group", "Subsector", "Indices"}]


    all_daily_records = []  # accumulates ALL records across all dates for daily dashboard

    for target_date in dates:
        date_str = target_date.strftime("%d %b %Y")
        sheet_name = safe_name(target_date.strftime("%d-%b-%Y"))

        if len(dates) > 1:
            print(f"\n-- {date_str} --")

        # Scrape results list for this date
        companies = scrape_results_list(session, target_date)
        if not companies:
            print(f"  No results for {date_str} - skipping.")
            continue

        print(f"  [{date_str}] Fetching {len(companies)} company pages...")
        records = enrich_all(session, companies, cache, date_str)
        save_cache(cache)

        df = pd.DataFrame(records)
        for c in num_cols:
            if c in df.columns:
                df[c] = pd.to_numeric(df[c], errors="coerce")
        
        if "Indices" in df.columns:
            counts = df["Indices"].value_counts()
            print(f"    Indices summary: {dict(counts)}")
        else:
            print("    CRITICAL: Indices column missing from DataFrame!")

        df_all = df.copy()
        df_1000 = filter_by_mcap(df)

        print(
            f"  Total: {len(df)} | MCap >= {MCAP_FLOOR} Cr: {len(df_1000)} | All companies: {len(df_all)} | Sectors: {df_all['Sector'].nunique()}")

        # Results_By_Date.xlsx
        wb_d = load_or_new(DATE_WB_PATH)
        if sheet_name in wb_d.sheetnames:
            del wb_d[sheet_name]
        ws = wb_d.create_sheet(sheet_name)
        write_date_sheet(ws, date_str, df_all)  # MCap filter is client-side in dashboard
        wb_d.save(DATE_WB_PATH)

        # Results_By_Sector.xlsx
        wb_s = load_or_new(SECTOR_WB_PATH)
        for sector in sorted(df_all["Sector"].dropna().unique()):
            df_new = df_all[df_all["Sector"] == sector].copy()
            sname = safe_name(sector)
            if sname not in wb_s.sheetnames:
                ws2 = wb_s.create_sheet(sname)
                write_sector_fresh(ws2, sector, df_new)
            else:
                ws2 = wb_s[sname]
                df_exist = read_sheet_df(ws2)
                df_comb = pd.concat([df_exist, df_new], ignore_index=True)
                df_comb = df_comb.drop_duplicates(
                    subset=["Company Name", "Quarter"], keep="last")
                for c in num_cols:
                    if c in df_comb.columns:
                        df_comb[c] = pd.to_numeric(df_comb[c], errors="coerce")
                rewrite_sector_data(ws2, df_comb)
        wb_s.save(SECTOR_WB_PATH)

        # Collect for daily dashboard (all companies — MCap floor applied client-side)
        all_daily_records.append(df_all)

    # -- After all dates processed --------------------------------------
    print(f"\n[3/4] Building Sector Summary & Dashboards...")

    # Reload sector workbook for summary (may have been updated across multiple dates)
    wb_s = load_or_new(SECTOR_WB_PATH)
    df_full = load_all_sector_data(wb_s)
    df_full_1000 = filter_by_mcap(df_full)
    df_sum, n500_avgs = build_sector_summary(wb_s, df_full_1000)
    wb_s.save(SECTOR_WB_PATH)
    print(f"  [OK] {SECTOR_WB_PATH}")

    print("[4/4] Writing dashboards...")
    build_html_dashboard(df_sum, n500_avgs, DASHBOARD_HTML)

    # Daily dashboard - covers all dates processed in this run
    if all_daily_records:
        df_daily = pd.concat(all_daily_records, ignore_index=True)
        build_daily_dashboard(df_daily, DAILY_DASHBOARD_HTML, date_label)
    else:
        print("  No data collected - Daily Dashboard skipped.")

    print(f"\n[OK] Done.  ({date_label})")
    if all_daily_records:
        df_daily_all = pd.concat(all_daily_records, ignore_index=True)
        print(f"  Companies in daily dashboard: {len(df_daily_all)}")
    print(
        f"  Sectors in summary: {df_sum['Sector'].nunique() if not df_sum.empty and 'Sector' in df_sum.columns else 0}")


if __name__ == "__main__":
    main()
