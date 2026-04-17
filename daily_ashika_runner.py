import os
import sys
import json
import subprocess
import pandas as pd
from pathlib import Path
from datetime import datetime, date

ROOT = Path(__file__).parent.absolute()
RESULTS_DIR = ROOT / "results_logic"
VOLUME_DIR = ROOT / "volume_logic"
MASTER_FILE = ROOT / "company_master.json"
BOARD_CSV = RESULTS_DIR / "Board_Meetings.csv"
RESULTS_XLS = RESULTS_DIR / "Results_By_Date.xlsx"

PYTHON = sys.executable

def log(msg):
    print(f"  [ASHIKA MASTER] {msg}")

def run_results_engine():
    log("Executing Results Engine...")
    script = RESULTS_DIR / "logic_results.py"
    subprocess.run([PYTHON, str(script)], cwd=RESULTS_DIR)

def sync_master_repository():
    log("Syncing Company Master Repository...")
    if not MASTER_FILE.exists():
        log("Master file missing, initializing...")
        from init_master import init_master
        init_master()

    with open(MASTER_FILE, "r") as f:
        master = json.load(f)

    # 0. Load Universe mapping
    univ_path = VOLUME_DIR / "screener_cache" / "bse2000_universe.json"
    bse_to_sym = {}
    if univ_path.exists():
        with open(univ_path, "r") as f:
            univ = json.load(f)
            bse_to_sym = {str(v).strip(): k for k, v in univ.get("bse_codes", {}).items() if v}

    # 1. Sync from Board Meetings (Upcoming)
    if BOARD_CSV.exists():
        log("Updating upcoming meeting dates...")
        try:
            df_board = pd.read_csv(BOARD_CSV)
            # Board CSV has "Security Code" (BSE) and "Purpose" and "Meeting Date"
            for _, row in df_board.iterrows():
                bse_code = str(row.get("Security Code", "")).strip()
                sym = bse_to_sym.get(bse_code)
                if sym and sym in master:
                    master[sym]["upcoming_meeting"] = str(row.get("Meeting Date", ""))
                    master[sym]["purpose"] = str(row.get("Purpose", ""))
        except Exception as e:
            log(f"Board sync error: {e}")

    # 2. Sync from Results Excel (Last Result)
    if RESULTS_XLS.exists():
        log("Updating last result dates...")
        try:
            from openpyxl import load_workbook
            wb = load_workbook(RESULTS_XLS, read_only=True)
            for sname in wb.sheetnames:
                # Sheet names like "14-Apr-2026"
                try:
                    dt_obj = datetime.strptime(sname.strip(), "%d-%b-%Y")
                    dt_str = dt_obj.strftime("%d %b")
                    ws = wb[sname]
                    # Row data has "Screener Ticker" or similar.
                    # As a shortcut, results_logic only processes BSE symbols currently
                    # but we can look for the name or ticker.
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if not row or len(row) < 10: continue
                        ticker = str(row[9]).strip().upper() # Screener Ticker
                        if ticker in master:
                            master[ticker]["last_result"] = dt_str
                except: continue
            wb.close()
        except Exception as e:
            log(f"Results sync error: {e}")

    # Save
    with open(MASTER_FILE, "w") as f:
        json.dump(master, f, indent=2)

def run_volume_engine():
    log("Executing Volume Engine...")
    script = VOLUME_DIR / "logic_volume.py"
    subprocess.run([PYTHON, str(script)], cwd=VOLUME_DIR)

def generate_gateway_hub():
    log("Generating Ashika Gateway Hub (index.html)...")
    # This will be implemented in Phase 3
    pass

def main():
    log("Starting Daily Ashika Pipeline...")
    
    # 1. Update Corporate Data
    run_results_engine()
    
    # 2. Sync Repository
    sync_master_repository()
    
    # 3. Update Volume Data (Enriched by Master)
    run_volume_engine()
    
    # 4. Generate Hub
    generate_gateway_hub()
    
    log("Ashika Pipeline Complete.")
    
    # Auto-open dashboard
    index_path = ROOT / "index.html"
    if index_path.exists():
        log(f"Opening dashboard: {index_path}")
        os.startfile(index_path)

if __name__ == "__main__":
    main()
